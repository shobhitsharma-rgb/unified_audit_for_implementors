import pandas as pd
import streamlit as st
import io
import re
import numpy as np
from datetime import datetime, date

def check_duplicate_columns(file_obj):
    """
    Checks for duplicate column headers in an uploaded file.
    Returns: list of duplicate column names, or None if no duplicates.
    """
    try:
        if file_obj.name.lower().endswith('.csv'):
            file_obj.seek(0)
            # Read only first row to see raw headers
            df_h = pd.read_csv(file_obj, header=None, nrows=1)
        else:
            file_obj.seek(0)
            df_h = pd.read_excel(file_obj, header=None, nrows=1)
        
        if df_h.empty: return None
        
        headers = [str(h).strip() for h in df_h.iloc[0].tolist() if pd.notna(h) and str(h).strip() != ""]
        
        seen = set()
        dupes = []
        for h in headers:
            if h in seen:
                if h not in dupes: dupes.append(h)
            seen.add(h)
        
        file_obj.seek(0) # Reset for next read
        return dupes if dupes else None
    except Exception:
        file_obj.seek(0)
        return None

def format_datetime_strings(df, columns):
    """
    Attempts to convert values in specified columns to MM/DD/YYYY format.
    Safely handles strings, NaNs, and various date formats.
    """
    import pandas as pd
    for col in columns:
        if col in df.columns:
            def _clean_date_val(val):
                if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() in ["nan", "nat"]:
                    return ""
                try:
                    # Use pd.to_datetime for robust parsing
                    d = pd.to_datetime(str(val).strip(), errors='coerce')
                    if pd.notna(d):
                        return d.strftime('%m/%d/%Y')
                except (ValueError, TypeError):
                    pass
                return str(val).strip()
            df[col] = df[col].apply(_clean_date_val)
    return df

# --- US State Name to Abbreviation ---
US_STATE_TO_ABBR = {
    'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR',
    'california': 'CA', 'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE',
    'florida': 'FL', 'georgia': 'GA', 'hawaii': 'HI', 'idaho': 'ID',
    'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA', 'kansas': 'KS',
    'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
    'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
    'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV',
    'new hampshire': 'NH', 'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY',
    'north carolina': 'NC', 'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK',
    'oregon': 'OR', 'pennsylvania': 'PA', 'rhode island': 'RI', 'south carolina': 'SC',
    'south dakota': 'SD', 'tennessee': 'TN', 'texas': 'TX', 'utah': 'UT',
    'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA', 'west virginia': 'WV',
    'wisconsin': 'WI', 'wyoming': 'WY', 'district of columbia': 'DC',
    'puerto rico': 'PR', 'guam': 'GU', 'virgin islands': 'VI',
    'american samoa': 'AS', 'northern mariana islands': 'MP',
}
# Common misspellings
US_STATE_TO_ABBR['forida'] = 'FL'
US_STATE_TO_ABBR['califronia'] = 'CA'
US_STATE_TO_ABBR['massachuetts'] = 'MA'
US_STATE_TO_ABBR['pennsylvannia'] = 'PA'
US_STATE_TO_ABBR['conneticut'] = 'CT'
US_STATE_TO_ABBR['tennesee'] = 'TN'
US_STATE_TO_ABBR['missisippi'] = 'MS'
US_STATE_TO_ABBR['flordia'] = 'FL'
US_STATE_TO_ABBR['goergia'] = 'GA'
US_STATE_TO_ABBR['viginia'] = 'VA'
US_STATE_TO_ABBR['west virgnia'] = 'WV'
US_STATE_TO_ABBR['minnesotta'] = 'MN'
US_STATE_TO_ABBR['louisianna'] = 'LA'

# Reverse map for validation (abbreviation -> True)
VALID_ABBRS = set(US_STATE_TO_ABBR.values())

def convert_state_to_abbreviation(df, column):
    """
    Converts full US state names in the given column to 2-letter abbreviations.
    Already-abbreviated values are uppercased. Non-state values are left untouched.
    """
    if column not in df.columns:
        return df
    
    def _convert(val):
        if pd.isna(val) or str(val).strip() == '':
            return ''
        val_clean = str(val).strip()
        # If already a valid 2-letter abbreviation (case-insensitive)
        if val_clean.upper() in VALID_ABBRS:
            return val_clean.upper()
        # Try full name lookup
        lookup = val_clean.lower().strip()
        if lookup in US_STATE_TO_ABBR:
            return US_STATE_TO_ABBR[lookup]
        # Not a recognized state — return as-is (could be a license number, etc.)
        return val_clean
    
    df[column] = df[column].apply(_convert)
    return df

# --- Hardcoded Mappings ---

# Map Uzio Raw Headers -> Internal Standard Names
UZIO_RAW_MAPPING = {
    'Employee ID*': 'Employee ID',
    'Employee First Name*': 'First Name',
    'Employee Last Name*': 'Last Name',
    'Employee Middle Initial': 'Middle Initial',
    'Employee Suffix': 'Suffix',
    'Employment Status*': 'Employment Status',
    'Date of Hire*': 'Hire Date',
    'Original DOH': 'Original Hire Date',
    'Termination Date': 'Termination Date',
    'Termination Reason': 'Termination Reason',
    'Employment Type*': 'Employment Type',
    'Pay Type*': 'Pay Type',
    'Annual Salary(Digits)**': 'Annual Salary',
    'Hourly Pay Rate**': 'Hourly Pay Rate',
    'Working Hours per Week(Digits)**': 'Working Hours',
    'Job Title': 'Job Title',
    'Department': 'Department',
    'Official Email*': 'Work Email',
    'Personal Email': 'Personal Email',
    'Phone Number(Digits)': 'Phone Number',
    'Employee SSN': 'SSN',
    'Employee Date of Birth*': 'DOB',
    'Employee Gender*': 'Gender',
    'Employee Tobacco usage in last 12 months': 'Tobacco User',
    'FLSA Classification': 'FLSA Classification',
    'Employee Address Line 1': 'Address Line 1',
    'Employee Address Line 2': 'Address Line 2',
    'City*': 'City',
    'Zipcode*': 'Zip',
    'State(Abbreviation)*': 'State',
    'Mailing Address Line 1': 'Mailing Address Line 1',
    'Mailing Address Line 2': 'Mailing Address Line 2',
    'Mailing City': 'Mailing City',
    'Mailing Zipcode': 'Mailing Zip',
    'Mailing State(Abbreviation)': 'Mailing State',
    'Reporting Manager ID': 'Reports To ID',
    'Work Location': 'Work Location',
    'License Number*': 'License Number',
    'License Expiration Date': 'License Expiration Date'
}

def read_uzio_raw_file(uploaded_file):
    """
    Reads the raw Uzio .xlsm export.
    Expects 'Employee Details' sheet.
    Headers are in Row 4 (Index 3).
    Renames columns to Internal Standard Names.
    """
    try:
        # Read Excel - header=3 means 4th row is header
        df = pd.read_excel(uploaded_file, sheet_name='Employee Details', header=3)
        
        # Strip whitespace and normalize columns for matching
        df.columns = [str(c).strip() for c in df.columns]
        
        # Use robust normalization for mapping
        norm_mapping = {norm_colname(k).casefold(): v for k, v in UZIO_RAW_MAPPING.items()}
        
        new_cols = []
        for col in df.columns:
            nc = norm_colname(col).casefold()
            if nc in norm_mapping:
                new_cols.append(norm_mapping[nc])
            else:
                new_cols.append(col)
        df.columns = new_cols
            
        # Ensure 'Employee ID' is string (remove decimals if any)
        if 'Employee ID' in df.columns:
            df['Employee ID'] = df['Employee ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
        print("Uzio Raw File Read Successfully.")
        return df

    except Exception as e:
        st.error(f"Error reading Uzio Raw File: {e}")
        return None

def extract_mappings_from_uzio(df_source, df_template, vendor_map):
    """
    Extracts Job Title and Work Location mappings from a pre-filled Uzio template.
    Matches employees between source and template using Employee ID.
    Returns: (job_mappings, loc_mappings) dictionaries.
    """
    job_mappings = {}
    loc_mappings = {}
    
    src_id_col = vendor_map.get('Employee ID')
    src_job_col = vendor_map.get('Job Title')
    src_loc_col = vendor_map.get('Work Location')
    
    if not src_id_col or src_id_col not in df_source.columns:
        return job_mappings, loc_mappings
        
    # Uzio headers (Standard names after read_uzio_raw_file)
    # Note: UZIO_RAW_MAPPING maps 'Employee ID*' to 'Employee ID'
    UZIO_ID = 'Employee ID'
    UZIO_JOB = 'Job Title'
    UZIO_LOC = 'Work Location'
    
    if UZIO_ID not in df_template.columns or UZIO_JOB not in df_template.columns or UZIO_LOC not in df_template.columns:
        # Try raw names just in case it wasn't processed
        if 'Employee ID*' in df_template.columns: UZIO_ID = 'Employee ID*'
        else: return job_mappings, loc_mappings
        
    # Build a lookup for Uzio records
    uzio_lookup = {}
    for _, t_row in df_template.iterrows():
        tid = str(t_row.get(UZIO_ID, "")).strip().replace(".0", "")
        if tid:
            uzio_lookup[tid] = {
                'job': str(t_row.get(UZIO_JOB, "")).strip(),
                'loc': str(t_row.get(UZIO_LOC, "")).strip()
            }
            
    # Iterate source and build mapping
    for _, s_row in df_source.iterrows():
        sid = str(s_row.get(src_id_col, "")).strip().replace(".0", "")
        if not sid or sid not in uzio_lookup:
            continue
            
        u_job = uzio_lookup[sid]['job']
        u_loc = uzio_lookup[sid]['loc']
        
        s_job = str(s_row.get(src_job_col, "")).strip() if src_job_col else ""
        s_loc = str(s_row.get(src_loc_col, "")).strip() if src_loc_col else ""
        
        if s_job and u_job and u_job.lower() not in ['nan', 'none', '']:
            job_mappings[s_job] = u_job
        if s_loc and u_loc and u_loc.lower() not in ['nan', 'none', '']:
            loc_mappings[s_loc] = u_loc
                
    return job_mappings, loc_mappings

def norm_col(c):
    """Normalize column names to be case-insensitive and stripped."""
    if c is None: return ""
    return str(c).strip().replace("\n", " ").strip()

def norm_colname(c: str) -> str:
    """Robust column normalization: handles newlines, special quotes, and bracketed suffixes."""
    if c is None:
        return ""
    c = str(c).replace("\n", " ").replace("\r", " ")
    c = c.replace("\u00A0", " ")
    c = c.replace("’", "'").replace("“", '"').replace("”", '"')
    # Remove bracketed suffixes like (Personal Profile) or (Employment Profile - Pay Rates)
    c = re.sub(r'\(.*?\)', '', c)
    c = re.sub(r"\s+", " ", c).strip()
    c = c.replace("*", "")
    c = c.strip('"').strip("'")
    return c

def norm_key_series(s: pd.Series) -> pd.Series:
    """Normalize a pandas Series of keys (like Employee IDs) using the central norm_id logic."""
    return s.apply(norm_id)

def norm_blank(x):
    """Normalize blank/NaN values to an empty string."""
    if x is None:
        return ""
    if isinstance(x, float) and np.isnan(x):
        return ""
    if isinstance(x, str) and x.strip().lower() in {"", "nan", "none", "null"}:
        return ""
    return x

def try_parse_date(x):
    """Attempt to parse various date formats into standard MM/DD/YYYY string."""
    x = norm_blank(x)
    if x == "":
        return ""
    if isinstance(x, (datetime, date, np.datetime64, pd.Timestamp)):
        ts = pd.to_datetime(x, errors='coerce')
        if pd.isna(ts):
            return ""
        return ts.strftime("%m/%d/%Y")
    if isinstance(x, str):
        s = x.strip()
        try:
            ts = pd.to_datetime(s, errors='coerce')
            if pd.isna(ts):
                return s  # return the original string if it can't be parsed
            return ts.strftime("%m/%d/%Y")
        except Exception:
            return s
    return str(x)

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    If multiple columns normalize to the same name (case-insensitive), keep only the first one.
    This prevents .loc[index, col] from returning a Series instead of a scalar.
    """
    df = df.copy()
    norm_cols = [norm_colname(c).casefold() for c in df.columns]
    seen = set()
    to_keep = []
    for i, nc in enumerate(norm_cols):
        if nc not in seen:
            seen.add(nc)
            to_keep.append(i)
    return df.iloc[:, to_keep]

def safe_val(df, idx, col):
    """Safely get a scalar value from a dataframe, even if duplicate columns somehow exist."""
    if idx is None or col not in df.columns:
        return ""
    val = df.loc[idx, col]
    if isinstance(val, pd.Series):
        return val.iloc[0]
    return val

def normalize_space_and_case(x):
    """Strip whitespace, collapse consecutive spaces, and convert to casefold."""
    x = norm_blank(x)
    if x == "":
        return ""
    s = str(x).strip()
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()

def find_col(df_cols, *candidate_names):
    """Find a column in a list of columns by matching against several candidates (case-insensitive)."""
    norm_map = {norm_colname(c).casefold(): c for c in df_cols}
    for cand in candidate_names:
        key = norm_colname(cand).casefold()
        if key in norm_map:
            return norm_map[key]
    return None

def as_float_or_none(x):
    """Convert a value to float or return None if not a number."""
    x = norm_blank(x)
    if x == "":
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        try:
            return float(x)
        except Exception:
            return None
    if isinstance(x, str):
        s = x.strip().replace(",", "").replace("$", "")
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None

# --- Job Titles that Uzio ALWAYS treats as Hourly/Non-Exempt ---
# Any salaried employee with one of these job titles must be flagged,
# and a blank-Pay-Type/blank-FLSA employee with one of these titles is
# force-set to Hourly + Non-Exempt regardless of source values.
HOURLY_ONLY_JOB_TITLES = {
    "driver",
    "lead driver",
    "walker",
    "helper",
    "driver-lite",
    "driver-step van",
    "driver-unscheduled",
    "ddu dedicated",
    "ddu shared",
    "delivery associate",
    "delivery associates",
    "driver -major appliance",
}

# Whole-word, case-insensitive regex over the title set above.
# Matches "Lead Driver" via "driver" but NOT "Drivership"; matches "Dog Walker"
# via "walker" but NOT "Sidewalker".
_HOURLY_ONLY_JOB_TITLE_REGEX = re.compile(
    r'\b(?:' + '|'.join(re.escape(t) for t in sorted(HOURLY_ONLY_JOB_TITLES, key=len, reverse=True)) + r')\b',
    re.IGNORECASE,
)


def is_hourly_only_job_title(jt_val: str) -> bool:
    """Return True if the job title contains any canonical hourly-only role
    name as a whole word (case-insensitive)."""
    if jt_val is None:
        return False
    s = str(jt_val).strip()
    if not s or s.lower() == "nan":
        return False
    return bool(_HOURLY_ONLY_JOB_TITLE_REGEX.search(s))

def norm_ssn_canonical(x):
    """Normalize SSN to 9 digits, no dashes, padded with zeros."""
    x = norm_blank(x)
    if x == "":
        return ""
    s = str(x).strip().replace("-", "").replace(" ", "")
    # Remove decimal if exists
    if s.endswith(".0"):
        s = s[:-2]
    # Keep only digits
    s = re.sub(r"\D", "", s)
    if not s:
        return ""
    return s.zfill(9)[-9:]

def norm_id(x):
    """Normalize Employee ID: strip, convert to string, and remove leading zeros."""
    if pd.isna(x) or x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0"): # Handle potential float conversions
        s = s[:-2]
    # Remove leading zeros unless it's just "0"
    if s != "0":
        s = s.lstrip("0")
    return s

def get_identity_match_map(df_uzio, df_vendor, uzio_id_col, vendor_id_col, uzio_ssn_col, vendor_ssn_col):
    """
    Returns mapping[uzio_id_norm] = vendor_id_norm for records that share the same SSN but have different IDs.
    This ensures that when auditing, records with different IDs but matching SSNs can be linked.
    """
    match_map = {}
    
    # Get all IDs present in each system, normalized
    uzio_ids = set(norm_id(x) for x in df_uzio[uzio_id_col] if pd.notna(x) and x != "")
    v_ids = set(norm_id(x) for x in df_vendor[vendor_id_col] if pd.notna(x) and x != "")
    
    # Identify IDs that don't have a direct match
    stray_uz = uzio_ids - v_ids
    stray_v = v_ids - uzio_ids
    
    if not stray_uz or not stray_v:
        return match_map
        
    # Build a lookup for stray vendor records by SSN
    v_ssn_lookup = {}
    for _, row in df_vendor.iterrows():
        vid = norm_id(row.get(vendor_id_col))
        if not vid or vid not in stray_v:
            continue
        ssn_val = norm_ssn_canonical(row.get(vendor_ssn_col))
        if ssn_val:
            # If multiple people have the same SSN (unlikely but possible), 
            # we only take the first one to avoid ambiguous matching
            if ssn_val not in v_ssn_lookup:
                v_ssn_lookup[ssn_val] = vid
                
    # Now check stray Uzio records against the SSN lookup
    for _, row in df_uzio.iterrows():
        uzid = norm_id(row.get(uzio_id_col))
        if not uzid or uzid not in stray_uz:
            continue
        ssn_val = norm_ssn_canonical(row.get(uzio_ssn_col))
        if ssn_val and ssn_val in v_ssn_lookup:
            vid = v_ssn_lookup[ssn_val]
            match_map[uzid] = vid
            
    return match_map

def detect_duplicate_ssns(df, id_col, ssn_col):
    """
    Returns a dictionary of SSNs mapping to a list of IDs sharing that SSN.
    Only returns SSNs that are associated with more than one unique ID.
    Excludes blank SSNs.
    """
    id_ssn_map = {}
    for _, row in df.iterrows():
        eid = norm_id(row.get(id_col, ""))
        ssn = norm_ssn_canonical(row.get(ssn_col, ""))
        if not ssn or not eid:
            continue
        if ssn not in id_ssn_map:
            id_ssn_map[ssn] = set()
        id_ssn_map[ssn].add(eid)
    
    # Filter for duplicates
    duplicates = {ssn: sorted(list(ids)) for ssn, ids in id_ssn_map.items() if len(ids) > 1}
    return duplicates

def clean_money_val(x):
    """Parse money/percentage strings to float. Returns original string if not a number."""
    if pd.isna(x) or x == "":
        return 0.0
    s = str(x).strip()
    s_clean = s.replace("$", "").replace("%", "").replace(",", "")
    s_clean = s_clean.replace("(", "-").replace(")", "") # Handle accounting negative
    try:
        return float(s_clean)
    except:
        return 0.0

def validate_source_data(df_source, resolved_field_map):
    """
    Comprehensive pre-generation sanity checks on the raw source data.
    Returns a dict:
      - 'hard_errors': pd.DataFrame of blocking issues (blank mandatory fields, bad zip, missing salary)
      - 'flsa_corrections': pd.DataFrame of FLSA auto-corrections made
      - 'email_fallbacks': pd.DataFrame of email fallback auto-fills made
    The caller should block generation if hard_errors is not empty.
    """
    hard_errors = []
    flsa_corrections = []
    flsa_blanks = []
    type_blanks = []
    intern_corrections = []
    email_fallbacks = []
    salaried_drivers = []
    anomalies = []
    inactive_statuses = []
    position_blanks = []
    dol_status_blanks = []
    smart_driver_fixes = []
    zip_fixes = []

    # Get column names 
    # Hourly Exempt / Salaried Non-Exempt flags
    
    # Resolve column references
    emp_id_col = resolved_field_map.get('Employee ID')
    status_col = resolved_field_map.get('Employment Status')
    type_col = resolved_field_map.get('Employment Type')
    pay_type_col = resolved_field_map.get('Pay Type')
    job_title_col = resolved_field_map.get('Job Title')
    dept_col = resolved_field_map.get('Department')
    location_col = resolved_field_map.get('Work Location')
    zip_col = resolved_field_map.get('Zip')
    salary_col = resolved_field_map.get('Annual Salary')
    flsa_col = resolved_field_map.get('FLSA Classification')
    work_email_col = resolved_field_map.get('Work Email')
    personal_email_col = resolved_field_map.get('Personal Email')
    hours_col = resolved_field_map.get('Working Hours')
    state_col = resolved_field_map.get('State')
    hire_date_col = resolved_field_map.get('Hire Date')
    term_date_col = resolved_field_map.get('Termination Date')
    first_name_col = resolved_field_map.get('First Name')
    last_name_col = resolved_field_map.get('Last Name')
    ssn_col = resolved_field_map.get('SSN')
    
    # --- PRE-SCAN for DUPLICATE SSNs ---
    duplicate_ssns = set()
    if ssn_col and ssn_col in df_source.columns and emp_id_col and emp_id_col in df_source.columns:
        # Group by SSN and identify those linked to more than one unique Employee ID
        # Note: We filter for valid SSNs (not blank) to find true duplicates
        valid_ssn_mask = df_source[ssn_col].notna() & (df_source[ssn_col].astype(str).str.strip() != "")
        if valid_ssn_mask.any():
            temp_df = df_source[valid_ssn_mask][[emp_id_col, ssn_col]].copy()
            temp_df[ssn_col] = temp_df[ssn_col].astype(str).str.strip()
            temp_df[emp_id_col] = temp_df[emp_id_col].astype(str).str.strip()
            
            ssn_counts = temp_df.groupby(ssn_col)[emp_id_col].nunique()
            duplicate_ssns = set(ssn_counts[ssn_counts > 1].index)
    
    def get_emp_ref(row, idx):
        ref = f"Row {idx+2}"
        if emp_id_col and emp_id_col in df_source.columns:
            eid = row.get(emp_id_col)
            if pd.notna(eid) and str(eid).strip():
                ref = str(eid).strip()
        return ref

    def get_emp_name(row):
        fname = ""
        lname = ""
        if first_name_col and first_name_col in df_source.columns:
            val = row.get(first_name_col)
            fname = str(val).strip() if pd.notna(val) else ""
        if last_name_col and last_name_col in df_source.columns:
            val = row.get(last_name_col)
            lname = str(val).strip() if pd.notna(val) else ""
        return f"{fname} {lname}".strip()
    
    for idx, row in df_source.iterrows():
        emp_ref = get_emp_ref(row, idx)
        
        # --- HARD STOP CHECKS ---
        missing = []
        
        if ssn_col and ssn_col in df_source.columns:
            ssn_val_raw = row.get(ssn_col)
            ssn_val = str(ssn_val_raw).strip() if pd.notna(ssn_val_raw) else ""
            if not ssn_val:
                missing.append("SSN (blank)")
            elif ssn_val in duplicate_ssns:
                missing.append(f"Duplicate SSN found across different IDs ({ssn_val})")

        # 1. Blank Employment Status & Broad Status Checks
        if status_col and status_col in df_source.columns:
            raw_status = row.get(status_col)
            status_val = str(raw_status).strip() if pd.notna(raw_status) else ""
            status_lower = status_val.lower()

            if not status_val:
                missing.append("Employment Status (blank)")
            else:
                # A. Check for non-standard statuses (neither Active nor Terminated)
                # We allow some common variants like 'A', 'T', 'Active', 'Terminated', 'Inactive'
                is_standard = status_lower in ['a', 't', 'i'] or any(s in status_lower for s in ['active', 'terminated', 'inactive'])
                
                if not is_standard or status_lower in ['a04l', 'a08v']:
                    missing.append(f"Non-standard Status ({status_val})")

                # B. Logic Check: Terminated/Inactive/Leave vs Termination Date
                is_term = status_lower in ['t'] or any(s in status_lower for s in ['terminated'])
                is_inactive = status_lower in ['i'] or any(s in status_lower for s in ['inactive'])
                is_leave = 'leave' in status_lower
                
                if is_leave or is_inactive:
                    if term_date_col and term_date_col in df_source.columns:
                        tdate = row.get(term_date_col)
                        if pd.isna(tdate) or str(tdate).strip() == "" or str(tdate).lower() == "nan":
                            # No term date -> Suggest Active and warn to exclude from payroll
                            anomalies.append({
                                'Employee ID': emp_ref,
                                'Name': get_emp_name(row),
                                'Issue': f"{status_val} Employee",
                                'Message': "Please make them excluded from payroll on Uzio"
                            })
                
                if is_term:
                    if term_date_col and term_date_col in df_source.columns:
                        tdate = row.get(term_date_col)
                        if pd.isna(tdate) or str(tdate).strip() == "" or str(tdate).lower() == "nan":
                            missing.append("Terminated but missing Termination Date")
        
        # 2. Blank Employment Type / DOL Status
        # Find DOL_Status column (it maps to Employment Type in Uzio)
        dol_col = next((c for c in df_source.columns if str(c).lower().strip().replace('_',' ') in ['dol status', 'dol_status', 'worker category description']), None)
        
        if type_col and type_col in df_source.columns:
            val = row.get(type_col)
            if pd.isna(val) or str(val).strip() == "":
                # Blank Employment Type is ALWAYS auto-filled to "Full Time" on
                # download (fix_dol_status, default-on). Report it as a fix, not a
                # hard error — never add it to `missing`/red box.
                dol_status_blanks.append({
                    'Employee ID': emp_ref,
                    'Current DOL Status': '(Blank)',
                    'Suggestion': 'Auto-fill to Full-Time'
                })
        
        # 3. Blank Pay Type
        pay_val = ""
        is_pay_type_blank = False
        if pay_type_col and pay_type_col in df_source.columns:
            pay_val = row.get(pay_type_col)
            if pd.isna(pay_val) or str(pay_val).strip() == "":
                is_pay_type_blank = True
                pay_val = ""
            else:
                pay_val = str(pay_val).strip().lower()
        
        # 4. Blank Job Title
        job_val_raw = row.get(job_title_col) if job_title_col and job_title_col in df_source.columns else ""
        job_val = str(job_val_raw).strip().lower() if pd.notna(job_val_raw) and str(job_val_raw).strip().lower() != "nan" else ""
        # Driver / hourly-only role detection — whole-word match against the
        # canonical title set (Driver, Walker, Helper, DDU Dedicated, etc.).
        is_driver = is_hourly_only_job_title(job_val)

        if job_title_col and job_title_col in df_source.columns:
            if not job_val:
                # Predict how the blank Job Title will actually be resolved on
                # download so the UI reports it accurately (and only flags the
                # genuinely-unfillable ones in the red box).
                dv = row.get(dept_col) if dept_col and dept_col in df_source.columns else None
                dept_val_jt = str(dv).strip() if pd.notna(dv) else ""
                fv = row.get(flsa_col) if flsa_col and flsa_col in df_source.columns else None
                flsa_str_jt = str(fv).strip().lower() if pd.notna(fv) else ""
                is_flsa_blank_jt = flsa_str_jt in ("", "nan")
                is_hourly_jt = ("hour" in pay_val) if pay_val else False
                # fix_blank_jt_to_driver only fires once FLSA is Non-Exempt; a blank
                # FLSA on an Hourly row is first filled to Non-Exempt, so treat that
                # as Non-Exempt-to-be.
                will_be_non_exempt = ("non-exempt" in flsa_str_jt) or ("non exempt" in flsa_str_jt) or is_flsa_blank_jt

                if dept_val_jt and dept_val_jt.lower() != "nan":
                    # Department column present (Paycom). If the department itself is a
                    # Driver/hourly-only role AND FLSA is blank, the smart-driver path
                    # handles it (and lists it) — don't double-list it here.
                    if not (is_hourly_only_job_title(dept_val_jt) and is_flsa_blank_jt):
                        position_blanks.append({
                            'Employee ID': emp_ref,
                            'Original Job Title': '(Blank)',
                            'Resolution': 'department',
                            'Suggestion': 'Fallback to Department'
                        })
                elif is_hourly_jt and will_be_non_exempt:
                    # No Department column (ADP): a blank title on a Non-Exempt Hourly
                    # employee is defaulted to "Driver" on download.
                    position_blanks.append({
                        'Employee ID': emp_ref,
                        'Original Job Title': '(Blank)',
                        'Resolution': 'driver-default',
                        'Suggestion': "Default to 'Driver' (Non-Exempt Hourly)"
                    })
                else:
                    # Cannot be auto-filled — genuinely needs a person to decide.
                    missing.append("Job Title (blank)")

        # 4c. FLSA Blank Check (Special logic for Drivers / hourly-only roles)
        is_flsa_blank = False
        if flsa_col and flsa_col in df_source.columns:
            flsa_val = row.get(flsa_col)
            if pd.isna(flsa_val) or str(flsa_val).strip() == "":
                is_flsa_blank = True

        if is_flsa_blank:
            # Smart Driver Check: Job Title OR Dept matches the hourly-only roster
            dept_val_raw = row.get(dept_col) if dept_col and dept_col in df_source.columns else ""
            dept_val = str(dept_val_raw).strip().lower() if pd.notna(dept_val_raw) else ""

            if is_driver:
                smart_driver_fixes.append({
                    'Employee ID': emp_ref,
                    'Position': job_val or '(Blank)',
                    'Dept': dept_val or '(Blank)',
                    'Issue': 'Blank FLSA (Driver/Hourly-only Position)',
                    'Suggestion': 'Auto-fix to Non-Exempt'
                })
            elif is_hourly_only_job_title(dept_val):
                smart_driver_fixes.append({
                    'Employee ID': emp_ref,
                    'Position': '(Blank)',
                    'Dept': dept_val,
                    'Issue': 'Blank Position & FLSA (Driver/Hourly-only Dept)',
                    'Suggestion': 'Auto-fill Position & set Non-Exempt'
                })
            else:
                flsa_blanks.append({
                    'Employee ID': emp_ref,
                    'Issue': 'Blank FLSA Classification',
                    'Suggestion': 'Auto-fill based on Pay Type'
                })
        
        if is_pay_type_blank and not (is_driver and is_flsa_blank):
            missing.append("Pay Type (blank)")
                
        # 4b. Blank Work Location
        if location_col and location_col in df_source.columns:
            val = row.get(location_col)
            if pd.isna(val) or str(val).strip() == "":
                missing.append("Work Location (blank)")
        
        # 5. Invalid Zip Code (must be 5 digits)
        if zip_col and zip_col in df_source.columns:
            zip_val = row.get(zip_col)
            if pd.isna(zip_val) or str(zip_val).strip() == "":
                # A blank zip stays blank on download — genuinely needs attention.
                missing.append("Zip Code (blank)")
            else:
                # Mirror the download-time _fix_zip_local transform exactly: keep
                # digits, pad a 4-digit zip with a leading zero, trim to 5. If the
                # result is a clean 5 digits the tool fixes it automatically; only
                # what's still wrong afterwards (1-3 digits) stays in the red box.
                import re
                raw = str(zip_val).strip()
                s = re.sub(r'[^0-9]', '', raw.split('.')[0].split('-')[0])
                if len(s) == 4:
                    s = '0' + s
                fixed = s[:5]
                if len(fixed) == 5:
                    if fixed != raw:
                        zip_fixes.append({
                            'Employee ID': emp_ref,
                            'Original Zip': raw,
                            'Fixed Zip': fixed
                        })
                    # else already a clean 5-digit zip — nothing to report
                else:
                    missing.append(f"Zip Code ('{zip_val}' is not 5 digits)")
        
        # 6. Salaried without Annual Salary
        if pay_val and ("salary" in pay_val or "salaried" in pay_val):
            if salary_col and salary_col in df_source.columns:
                sal_val = row.get(salary_col)
                if pd.isna(sal_val) or str(sal_val).strip() == "" or str(sal_val).strip() == "0":
                    missing.append("Annual Salary (required for Salaried)")
        
        # 6b. Salaried Hourly-Only Job Title Exception — always check, irrespective of salary value
        if pay_val and ("salary" in pay_val or "salaried" in pay_val):
            if job_title_col and job_title_col in df_source.columns:
                jt_raw = row.get(job_title_col)
                jt_val = str(jt_raw).strip().lower() if pd.notna(jt_raw) else ""
                if jt_val and jt_val != "nan":
                    if is_hourly_only_job_title(jt_val):
                        missing.append(f"Salaried Hourly-Only Exception (Job Title '{str(jt_raw).strip()}' must be Hourly/Non-Exempt in Uzio)")
                        salaried_drivers.append({
                            'Employee ID': emp_ref,
                            'Job Title': str(jt_raw).strip(),
                            'Pay Type': str(row.get(pay_type_col, '')).strip()
                        })
        
        # 7. Working Hours — no check. They are forced to 0 for every employee
        # at download time, so a blank or non-zero source value is not an issue.

        # 8. State must be 2-character abbreviation (not full name)
        if state_col and state_col in df_source.columns:
            state_val = row.get(state_col)
            if pd.notna(state_val) and str(state_val).strip():
                sv = str(state_val).strip()
                if len(sv) > 2:
                    missing.append(f"State ('{sv}' is full name, need 2-char abbreviation)")
                    
        # 9. Termination Date vs Hire Date validity
        if hire_date_col and hire_date_col in df_source.columns and term_date_col and term_date_col in df_source.columns:
            hire_val = row.get(hire_date_col)
            term_val = row.get(term_date_col)
            
            # Check if both dates are present (if either exists, both must be valid relative to each other)
            if pd.notna(hire_val) and str(hire_val).strip() != "" and pd.notna(term_val) and str(term_val).strip() != "":
                # Attempt to parse both dates
                try:
                    # Using pd.to_datetime with errors='coerce' to safely parse strings to datetimes
                    parsed_hire = pd.to_datetime(hire_val, errors='coerce')
                    parsed_term = pd.to_datetime(term_val, errors='coerce')
                    
                    if pd.notna(parsed_hire) and pd.notna(parsed_term):
                        if parsed_term < parsed_hire:
                            missing.append(f"Date of termination ({parsed_term.strftime('%Y-%m-%d')}) predates date of hire ({parsed_hire.strftime('%Y-%m-%d')})")
                except Exception:
                    pass # Ignore if dates are malformed, we just won't flag this specific error
        
        # --- Special Character Check & Fix (Emergency Contact & Relationship) ---
        emergency_cols = [c for c in df_source.columns if 'emergency' in str(c).lower()]
        for ec in emergency_cols:
            ec_raw = row.get(ec)
            val = str(ec_raw).strip()
            if pd.notna(ec_raw) and val and val.lower() != "nan":
                # Special Rule: Fiancée -> Fiancee (Handle accents/special chars)
                if val.lower().startswith("fian"):
                    # We will log this as a correction later, but for now we skip the error
                    val = "Fiancee"
                
                # Reverted to strict check: Allow alphanumeric, spaces, hyphens, and apostrophes only
                if not re.match(r"^[A-Za-z0-9\s\-\']+$", val):
                    missing.append(f"Special characters in {ec} ('{val}')")

        if missing:
            hard_errors.append({
                'Employee ID': emp_ref,
                'Name': get_emp_name(row),
                'Issue': ", ".join(missing)
            })
        
        # --- SOFT CHECKS (detection only — source FLSA is preserved as-is;
        # see generate_uzio_template for the actual fill behavior) ---

        # 7. FLSA Mismatch (flag only — does NOT overwrite source FLSA)
        if flsa_col and flsa_col in df_source.columns and pay_type_col and pay_type_col in df_source.columns:
            flsa_val = row.get(flsa_col)
            if pd.notna(flsa_val) and str(flsa_val).strip():
                flsa_str = str(flsa_val).strip().lower()

                if pay_val and ("hourly" in pay_val or "hour" in pay_val):
                    if "exempt" in flsa_str and "non" not in flsa_str:
                        # Hourly + Exempt mismatch — flag only, source preserved
                        flsa_corrections.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Pay Type': str(row.get(pay_type_col, '')).strip(),
                            'Original FLSA': str(flsa_val).strip(),
                            'Suggested FLSA (Source Preserved)': 'Non-Exempt'
                        })
                        anomalies.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Issue': f"Hourly Exempt (Pay Type: {str(row.get(pay_type_col, '')).strip()}, FLSA: {str(flsa_val).strip()})"
                        })
                elif pay_val and ("salary" in pay_val or "salaried" in pay_val):
                    if "non" in flsa_str and "exempt" in flsa_str:
                        # Salaried + Non-Exempt mismatch — flag only, source preserved
                        flsa_corrections.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Pay Type': str(row.get(pay_type_col, '')).strip(),
                            'Original FLSA': str(flsa_val).strip(),
                            'Suggested FLSA (Source Preserved)': 'Exempt'
                        })
                        anomalies.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Issue': f"Salaried Non-Exempt (Pay Type: {str(row.get(pay_type_col, '')).strip()}, FLSA: {str(flsa_val).strip()})"
                        })
            else:
                # FLSA is blank — soft flag
                if pay_val and ("hourly" in pay_val or "hour" in pay_val or "salary" in pay_val or "salaried" in pay_val):
                    flsa_blanks.append({
                        'Employee ID': emp_ref,
                        'Name': get_emp_name(row),
                        'Pay Type': str(row.get(pay_type_col, '')).strip(),
                        'FLSA Classification': '(blank)'
                    })
        
        # 8. Blank Work Email → fill with Personal Email
        if work_email_col and work_email_col in df_source.columns:
            we_val = row.get(work_email_col)
            if pd.isna(we_val) or str(we_val).strip() == "":
                if personal_email_col and personal_email_col in df_source.columns:
                    pe_val = row.get(personal_email_col)
                    if pd.notna(pe_val) and str(pe_val).strip():
                        email_fallbacks.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Personal Email Used': str(pe_val).strip()
                        })
        
        # 9. Intern in Worker Category → auto-correct to Part Time
        if type_col and type_col in df_source.columns:
            type_val = row.get(type_col)
            if pd.notna(type_val) and 'intern' in str(type_val).strip().lower():
                intern_corrections.append({
                    'Employee ID': emp_ref,
                    'Name': get_emp_name(row),
                    'Original Employment Type': str(type_val).strip(),
                    'Corrected Employment Type': 'Part Time'
                })
                
        # 10. Inactive Status tracking
        if status_col and status_col in df_source.columns:
            emp_status_val = str(row.get(status_col)).strip().lower()
            if emp_status_val == "inactive":
                inactive_statuses.append({
                    'Employee ID': emp_ref,
                    'Name': get_emp_name(row),
                    'Original Status': str(row.get(status_col)).strip(),
                    'Has Termination Date': 'Yes' if (term_date_col and term_date_col in df_source.columns and pd.notna(row.get(term_date_col)) and str(row.get(term_date_col)).strip() != "") else 'No'
                })
                
        # 11. (removed) Blank Position/Job Title tracking — this used to append a
        # second, differently-shaped row to position_blanks for every blank job
        # title, emitting a misleading "filled from department" / "Not available
        # (Department missing)" suggestion even for ADP, which has no department
        # column. Section 4 above now classifies blank job titles accurately
        # (department / driver-default / unresolved); this duplicate is dropped.

        # 12. DOL_Status tracking (Paycom primarily, but safe to check if mapped)
        dol_col = None
        for cand in ['dol_status', 'dol status']:
            cand_col = next((c for c in df_source.columns if str(c).lower().strip() == cand), None)
            if cand_col:
                dol_col = cand_col
                break
                
        if dol_col:
            dol_val = row.get(dol_col)
            if pd.isna(dol_val) or str(dol_val).strip() == "":
                emp_status_str = str(row.get(status_col)).strip().lower() if status_col and status_col in df_source.columns else ""
                if "term" not in emp_status_str and emp_status_str != "inactive":
                    dol_status_blanks.append({
                        'Employee ID': emp_ref,
                        'Name': get_emp_name(row),
                        'Status': 'Active (DOL_Status blank)'
                    })
    
    # 13. Create detailed summary for UI Dashboard
    error_summary = {
        'Missing Info': [],
        'Date & Status Logic': [],
        'Contact Formatting': []
    }
    
    for err in hard_errors:
        issue = err['Issue']
        eid = err['Employee ID']
        name = err.get('Name', '')
        ref = f"{eid} ({name})" if name else eid
        
        # Categorize
        is_date_status = any(m in issue for m in ['predates date of hire', 'Terminated/Inactive but missing', 'Non-standard Status'])
        is_contact = 'Special characters' in issue
        
        if is_date_status:
            error_summary['Date & Status Logic'].append(ref)
        elif is_contact:
            error_summary['Contact Formatting'].append(ref)
        else:
            # Everything else is missing info or basic formatting (SSN, Job Title, Zip, etc)
            error_summary['Missing Info'].append(ref)

    return {
        'hard_errors': pd.DataFrame(hard_errors),
        'flsa_corrections': pd.DataFrame(flsa_corrections),
        'flsa_blanks': pd.DataFrame(flsa_blanks),
        'type_blanks': pd.DataFrame(type_blanks),
        'intern_corrections': pd.DataFrame(intern_corrections),
        'email_fallbacks': pd.DataFrame(email_fallbacks),
        'salaried_drivers': pd.DataFrame(salaried_drivers).assign(Name=lambda d: d['Employee ID'].map(lambda x: next((e['Name'] for e in hard_errors if e['Employee ID'] == x), "")) if not d.empty else ""),
        'anomalies': pd.DataFrame(anomalies),
        'inactive_statuses': pd.DataFrame(inactive_statuses),
        'position_blanks': pd.DataFrame(position_blanks),
        'dol_status_blanks': pd.DataFrame(dol_status_blanks),
        'smart_driver_fixes': pd.DataFrame(smart_driver_fixes),
        'zip_fixes': pd.DataFrame(zip_fixes),
        'error_summary': error_summary
    }

def generate_uzio_template(df_source, vendor_field_map, fix_options=None):
    """
    Generate an Uzio Census Template DataFrame from a source DataFrame.
    """
    
    # Create an empty dataframe with Uzio headers
    uzio_headers = list(UZIO_RAW_MAPPING.keys())
    df_uzio = pd.DataFrame(columns=uzio_headers)
    
    # Iterate through each Uzio expected header
    for uzio_header, std_name in UZIO_RAW_MAPPING.items():
        # Job Title / Department / Work Location: populate from source so the
        # driver-mask check below can see real values. The Streamlit caller
        # overwrites these with the user-mapping dict after this function
        # returns (apps/{adp,paycom}/census_generator.py), so the final output
        # still respects the user's mapping — but driver detection now works.
        if std_name in ['Job Title', 'Department', 'Work Location']:
            vendor_col = vendor_field_map.get(std_name)
            if vendor_col and vendor_col in df_source.columns:
                df_uzio[uzio_header] = df_source[vendor_col].fillna("").astype(str).str.strip().values
            else:
                df_uzio[uzio_header] = ""
            continue
            
        vendor_col = vendor_field_map.get(std_name)
        if vendor_col and vendor_col in df_source.columns:
            # We have a direct mapping
            series = df_source[vendor_col].copy()
            
            # Apply formatting rules
            if std_name == 'Middle Initial':
                series = series.apply(lambda x: str(x).strip()[0] if pd.notna(x) and str(x).strip() else "")
            elif std_name in ['Hire Date', 'Original Hire Date', 'Termination Date', 'DOB']:
                def format_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    try:
                        dt = pd.to_datetime(str(d).strip(), errors='coerce')
                        if pd.isna(dt): return str(d).strip()
                        return dt.strftime('%d/%m/%Y')
                    except:
                        return str(d).strip()
                series = series.apply(format_date)
            elif std_name == 'License Expiration Date':
                def format_license_exp_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    d_str = str(d).strip()
                    # Never allow placeholder invalid dates
                    if '00/00/0000' in d_str or d_str in ('0', '00', '0000'): return ""
                    try:
                        dt = pd.to_datetime(d_str, errors='coerce')
                        if pd.isna(dt): return ""
                        return dt.strftime('%m/%d/%Y')  # Uzio wants MM/DD/YYYY
                    except:
                        return ""
                series = series.apply(format_license_exp_date)
            elif std_name == 'SSN':
                series = series.apply(lambda x: str(x).replace("-", "").strip() if pd.notna(x) else "")
            elif std_name == 'Gender':
                def format_gender(g):
                    if pd.isna(g) or str(g).strip() == "": return ""
                    g_str = str(g).strip().lower()
                    if g_str.startswith('m'): return "Male"
                    if g_str.startswith('f'): return "Female"
                    return ""
                series = series.apply(format_gender)
            elif std_name == 'Employment Status':
                def format_status(row):
                    x = row[vendor_col]
                    if pd.isna(x): return ""
                    s = str(x).strip().lower()
                    if not s: return ""
                    
                    if fix_options and fix_options.get('fix_status', False):
                        if 'not hired' in s: return 'EXCLUDE'
                        if 'leave' in s: return 'ACTIVE'
                        if 'term' in s: return 'TERMINATED'
                        if 'active' in s: return 'ACTIVE'
                    
                    if fix_options and fix_options.get('fix_inactive', False):
                        if 'inactive' in s:
                            # Only Terminate if Termination Date exists
                            term_col = vendor_field_map.get('Termination Date')
                            if term_col and pd.notna(row.get(term_col)) and str(row.get(term_col)).strip() != "":
                                return 'TERMINATED'
                            return 'ACTIVE' # Default inactive to active if no term date
                    elif 'inactive' in s:
                        return 'INACTIVE' # Preserve original if not fixed
                    
                    return str(x).strip().upper()
                series = df_source.apply(format_status, axis=1)
            elif std_name in ['Zip', 'Mailing Zip']:
                def format_zip(z):
                    if pd.isna(z) or str(z).strip() == "": return ""
                    
                    if fix_options and fix_options.get('fix_zip', False):
                        # Keep digits only
                        import re
                        # Trim after hyphen or decimal (user request)
                        s = str(z).split('.')[0].split('-')[0]
                        z_clean = re.sub(r'\D', '', s.strip())
                        if not z_clean: return ""
                        # Pad to 5 or truncate to 5
                        if len(z_clean) == 4:
                            return '0' + z_clean
                        else:
                            return z_clean[:5]
                    else:
                        # Return as-is if no fix requested
                        return str(z).strip()
                series = series.apply(format_zip)
            elif std_name == 'Employment Type':
                def format_emp_type(et):
                    if pd.isna(et) or str(et).strip() == "": return ""
                    et_str = str(et).strip().lower()
                    
                    if fix_options and fix_options.get('fix_type', False):
                        if 'full' in et_str: return 'Full Time'
                        if 'part' in et_str: return 'Part Time'
                        if 'season' in et_str: return 'Seasonal'
                        if 'other' in et_str: return 'Other'
                        if 'intern' in et_str: return 'Part Time'
                    
                    return str(et).strip()
                series = series.apply(format_emp_type)
            elif std_name == 'Termination Reason':
                def format_term_reason(tr):
                    if pd.isna(tr) or str(tr).strip() == "": return ""
                    tr_str = str(tr).strip().lower()
                    
                    if "involuntary" in tr_str or "invluntary" in tr_str:
                        return "Involuntary Termination of Employment"
                    if "voluntary" in tr_str or "quit" in tr_str:
                        return "Voluntary Termination of Employment"
                    if "death" in tr_str:
                        return "Death"
                    if "retire" in tr_str:
                        return "Retirement"
                    if "disability" in tr_str:
                        return "Permanent Disability"
                    if "transfer" in tr_str:
                        return "Transfer"
                    
                    # Anything else that is not blank gets 'Other'
                    return "Other"
                series = series.apply(format_term_reason)
            # We port the data
            df_uzio[uzio_header] = series
        else:
            df_uzio[uzio_header] = ""
    # Initialize log tracking
    fix_logs = []
    
    # Filter out excluded employees (e.g., 'not hired')
    if 'Employment Status*' in df_uzio.columns:
        df_uzio = df_uzio[df_uzio['Employment Status*'] != 'EXCLUDE'].copy()
        
    # Helper to get employee ID for logging
    emp_ids = df_uzio['Employee ID*'] if 'Employee ID*' in df_uzio.columns else df_uzio.index
        
    # Apply Work Email Fallback (Optional)
    if fix_options and fix_options.get('fix_emails', False):
        if 'Official Email*' in df_uzio.columns and 'Personal Email' in df_uzio.columns:
            # Fill missing Work Emails with Personal Email
            missing_work_mask = df_uzio['Official Email*'].isna() | (df_uzio['Official Email*'].astype(str).str.strip() == "")
            has_personal_mask = df_uzio['Personal Email'].notna() & (df_uzio['Personal Email'].astype(str).str.strip() != "")
            combined_mask = missing_work_mask & has_personal_mask
            
            for idx in df_uzio[combined_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx],
                    "Field Fixed": "Official Email*",
                    "Original Value": "(Blank)",
                    "New Value": df_uzio.loc[idx, 'Personal Email'],
                    "Fix Applied": "Fallback to Personal Email"
                })
                
            df_uzio.loc[combined_mask, 'Official Email*'] = df_uzio.loc[combined_mask, 'Personal Email']

    # Apply Position Auto-Fill (Optional - primarily Paycom)
    if fix_options and fix_options.get('fix_position', False):
        if 'Job Title' in df_uzio.columns:
            # Try to find department_desc column or similar for position fallback (Description only)
            dept_desc_col = next((c for c in df_source.columns if str(c).lower().strip().replace(' ','_') == 'department_desc' or str(c).lower().strip() == 'department_description'), None)
                    
            if dept_desc_col:
                missing_job_mask = df_uzio['Job Title'].isna() | (df_uzio['Job Title'].astype(str).str.strip() == "")
                has_dept_mask = df_source[dept_desc_col].notna() & (df_source[dept_desc_col].astype(str).str.strip() != "")
                combined_mask = missing_job_mask & has_dept_mask
                
                for idx in df_uzio[combined_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "Job Title",
                        "Original Value": "(Blank)",
                        "New Value": df_source.loc[idx, dept_desc_col],
                        "Fix Applied": "Fallback to Department"
                    })
                    
                df_uzio.loc[combined_mask, 'Job Title'] = df_source.loc[combined_mask, dept_desc_col]

    # Apply DOL_Status Auto-Fill (Optional - primarily Paycom)
    if fix_options and fix_options.get('fix_dol_status', False):
        dol_col = None
        for cand in ['dol_status', 'dol status', 'worker category description']:
            cand_col = next((c for c in df_source.columns if str(c).lower().strip().replace('_',' ') == cand), None)
            if cand_col:
                dol_col = cand_col
                break

        if dol_col and 'Employment Type*' in df_uzio.columns:
            # Mask for ALL employees with blank DOL_Status
            blank_dol_mask = df_source[dol_col].isna() | (df_source[dol_col].astype(str).str.strip() == "")
            combined_mask = blank_dol_mask
            
            for idx in df_uzio[combined_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx],
                    "Field Fixed": "Employment Type*",
                    "Original Value": "(Blank)",
                    "New Value": "Full Time",
                    "Fix Applied": "Default blank to Full Time"
                })
            
            # Apply the fix: set Employment Type to 'Full Time'
            df_uzio.loc[combined_mask, 'Employment Type*'] = "Full Time"

    # --- License Rules (Optional) ---
    if fix_options and fix_options.get('fix_license', False):
        # Rule 1: Never allow License Expiration Date if License Number is blank
        # Rule 2: Never allow 00/00/0000 in License Expiration Date
        lic_num_col = 'License Number*'
        lic_exp_col = 'License Expiration Date'
        if lic_exp_col in df_uzio.columns:
            # Clear 00/00/0000 or similar invalid placeholders
            bad_exp_mask = df_uzio[lic_exp_col].astype(str).str.strip().isin(['00/00/0000', '0', '00', '0000', 'nan', 'NaT', ''])
            
            for idx in df_uzio[bad_exp_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx],
                    "Field Fixed": "License Expiration Date",
                    "Original Value": df_uzio.loc[idx, lic_exp_col],
                    "New Value": "(Blank)",
                    "Fix Applied": "Cleared Invalid Date Placeholder"
                })
                
            df_uzio.loc[bad_exp_mask, lic_exp_col] = ""
            
            # Clear expiration date if no license number
            if lic_num_col in df_uzio.columns:
                no_license_mask = df_uzio[lic_num_col].isna() | (df_uzio[lic_num_col].astype(str).str.strip() == "") | (df_uzio[lic_num_col].astype(str).str.strip() == 'nan')
                exp_not_blank = df_uzio[lic_exp_col].astype(str).str.strip() != ""
                combined_mask = no_license_mask & exp_not_blank
                
                for idx in df_uzio[combined_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "License Expiration Date",
                        "Original Value": df_uzio.loc[idx, lic_exp_col],
                        "New Value": "(Blank)",
                        "Fix Applied": "Cleared Date due to missing License Number"
                    })
                    
                df_uzio.loc[no_license_mask, lic_exp_col] = ""


    # Apply Pay Type rules
    if 'Pay Type*' in df_uzio.columns:
        # Rule 1 (highest precedence): If Job Title matches the hourly-only
        # roster (Driver, Walker, Helper, DDU Dedicated, etc. — whole-word,
        # case-insensitive), force Pay Type = Hourly and FLSA = Non-Exempt
        # regardless of source values.
        if 'Job Title' in df_uzio.columns:
            driver_mask = df_uzio['Job Title'].apply(is_hourly_only_job_title)

            pt_to_fix = driver_mask & ((df_uzio['Pay Type*'].astype(str).str.lower().str.strip() != 'hourly') | df_uzio['Pay Type*'].isna() | (df_uzio['Pay Type*'] == ""))
            for idx in df_uzio[pt_to_fix].index:
                fix_logs.append({
                    "Employee": emp_ids[idx],
                    "Field Fixed": "Pay Type*",
                    "Original Value": df_uzio.loc[idx, 'Pay Type*'] if pd.notna(df_uzio.loc[idx, 'Pay Type*']) and str(df_uzio.loc[idx, 'Pay Type*']).strip() else "(Blank)",
                    "New Value": "Hourly",
                    "Fix Applied": "Forced Hourly for Driver/Hourly-only Position"
                })
            df_uzio.loc[driver_mask, 'Pay Type*'] = "Hourly"

            if 'FLSA Classification' in df_uzio.columns:
                flsa_to_fix = driver_mask & ((df_uzio['FLSA Classification'].astype(str).str.lower().str.strip() != 'non-exempt') | df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'] == ""))
                for idx in df_uzio[flsa_to_fix].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "FLSA Classification",
                        "Original Value": df_uzio.loc[idx, 'FLSA Classification'] if pd.notna(df_uzio.loc[idx, 'FLSA Classification']) and str(df_uzio.loc[idx, 'FLSA Classification']).strip() else "(Blank)",
                        "New Value": "Non-Exempt",
                        "Fix Applied": "Forced Non-Exempt for Driver/Hourly-only Position"
                    })
                df_uzio.loc[driver_mask, 'FLSA Classification'] = "Non-Exempt"
        else:
            driver_mask = pd.Series(False, index=df_uzio.index)

        pay_type_series = df_uzio['Pay Type*'].astype(str).str.lower().str.strip()

        # Hourly logic (normalize Pay Type label; clear Annual Salary)
        hourly_mask = pay_type_series.str.contains('hour', na=False)
        df_uzio.loc[hourly_mask, 'Pay Type*'] = "Hourly"
        if 'Annual Salary(Digits)**' in df_uzio.columns:
            df_uzio.loc[hourly_mask, 'Annual Salary(Digits)**'] = ""

        # Salaried logic (normalize Pay Type label; clear Hourly Rate / Working Hours)
        salary_mask = pay_type_series.str.contains('salar', na=False)
        df_uzio.loc[salary_mask, 'Pay Type*'] = "Salaried"
        if 'Hourly Pay Rate**' in df_uzio.columns:
            df_uzio.loc[salary_mask, 'Hourly Pay Rate**'] = 0
        if 'Working Hours per Week(Digits)**' in df_uzio.columns:
            df_uzio.loc[salary_mask, 'Working Hours per Week(Digits)**'] = ""

        # Rules 2-4: FLSA fill-by-Pay-Type — BLANKS ONLY, never overwrite a
        # source FLSA value. Driver rule (above) already covered its rows.
        if fix_options and fix_options.get('fix_flsa', False):
            if 'FLSA Classification' in df_uzio.columns:
                blank_flsa_mask = df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'].astype(str).str.strip() == "") | (df_uzio['FLSA Classification'].astype(str).str.strip().str.lower() == "nan")

                # Rule 3: blank FLSA + Hourly + not Driver → Non-Exempt
                hourly_fill_mask = blank_flsa_mask & hourly_mask & ~driver_mask
                for idx in df_uzio[hourly_fill_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "FLSA Classification",
                        "Original Value": "(Blank)",
                        "New Value": "Non-Exempt",
                        "Fix Applied": "Filled blank FLSA based on Hourly Pay Type"
                    })
                df_uzio.loc[hourly_fill_mask, 'FLSA Classification'] = "Non-Exempt"

                # Rule 2: blank FLSA + Salaried + not Driver → Exempt
                salary_fill_mask = blank_flsa_mask & salary_mask & ~driver_mask
                for idx in df_uzio[salary_fill_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "FLSA Classification",
                        "Original Value": "(Blank)",
                        "New Value": "Exempt",
                        "Fix Applied": "Filled blank FLSA based on Salaried Pay Type"
                    })
                df_uzio.loc[salary_fill_mask, 'FLSA Classification'] = "Exempt"

                # Rule 4: still-blank FLSA + not Driver + Pay Type also blank/unknown
                # → cannot determine. Leave blank, surface in change log.
                still_blank_mask = (df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'].astype(str).str.strip() == "")) & ~driver_mask
                for idx in df_uzio[still_blank_mask].index:
                    pt_cur = df_uzio.loc[idx, 'Pay Type*'] if pd.notna(df_uzio.loc[idx, 'Pay Type*']) and str(df_uzio.loc[idx, 'Pay Type*']).strip() else "(Blank)"
                    fix_logs.append({
                        "Employee": emp_ids[idx],
                        "Field Fixed": "FLSA Classification",
                        "Original Value": "(Blank)",
                        "New Value": "(Blank — Not Filled)",
                        "Fix Applied": f"Cannot derive FLSA — source FLSA is blank, Job Title is not in Driver/Hourly-only list, and Pay Type is '{pt_cur}'. Manual review required."
                    })
                # Intentionally do NOT default to Non-Exempt — leave blank.

    # Final step: Attach logs
    df_uzio.attrs['fix_logs'] = pd.DataFrame(fix_logs) if fix_logs else pd.DataFrame(columns=["Employee", "Field Fixed", "Original Value", "New Value", "Fix Applied"])

    return df_uzio

def inject_into_uzio_template(df_uzio, template_path="templates/Uzio_Census_Template.xlsm"):
    """
    Injects a formatted Uzio DataFrame into the standard Uzio .xlsm template.
    Preserves all sheets, instructions, and headers.
    Dynamically finds the row containing 'Employee First Name*' and starts data on the next row.
    """
    import openpyxl
    import os
    import re
    
    if isinstance(template_path, str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found at {template_path}")
        
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Employee Details']
    
    # Dynamically find the header row
    header_row = 4 # Fallback
    headers_in_template = {}
    
    for r in range(1, 10): # Search first 10 rows
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and re.sub(r'\s+', ' ', str(val)).strip() == 'Employee First Name*':
                header_row = r
                break
        if header_row == r:
            break
            
    # Map column headers
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            # Normalize to handle templates with embedded newlines like 'Employment\nStatus*'
            norm_val = re.sub(r'\s+', ' ', str(val)).strip()
            headers_in_template[norm_val] = col_idx

    # Write data starting at the row after the headers
    start_row = header_row + 1

    for row_idx, row_data in df_uzio.iterrows():
        excel_row = start_row + row_idx
        for col_name in df_uzio.columns:
            c_name_strip = re.sub(r'\s+', ' ', str(col_name)).strip()
            if c_name_strip in headers_in_template:
                col_idx = headers_in_template[c_name_strip]
                val = row_data[col_name]
                if pd.notna(val) and val != "":
                    ws.cell(row=excel_row, column=col_idx, value=val)
                    
    return wb

def validate_uzio_data(df_uzio):
    """
    Validates required fields for Uzio Census.
    Returns a DataFrame containing Employee ID and the list of missing fields.
    Fields checked: Pay Type*, Employment Status*, Job Title, Work Location.
    """
    errors = []
    
    # Identify expected column names from UZIO_RAW_MAPPING vs what's in df_uzio
    # Or just use the exact Uzio headers if df_uzio has them
    emp_id_col = 'Employee ID*' if 'Employee ID*' in df_uzio.columns else 'Employee ID'
    
    for idx, row in df_uzio.iterrows():
        emp_id = row.get(emp_id_col, f"Row {idx+1}")
        if pd.isna(emp_id) or str(emp_id).strip() == "":
            emp_id = f"Row {idx+1}"
            
        missing_fields = []
        
        # Check Pay Type
        val_pt = row.get('Pay Type*')
        if pd.isna(val_pt) or str(val_pt).strip() == "":
            missing_fields.append("Pay Type")
            
        # Check Employment Status
        val_es = row.get('Employment Status*')
        if pd.isna(val_es) or str(val_es).strip() == "":
            missing_fields.append("Employment Status")
            
        # Check Job Title
        val_jt = row.get('Job Title')
        invalid_jt = False
        allowed_titles = [
            'dsp owner', 'operations manager', 'operations lead', 'fleet manager', 
            'safety manager', 'performance manager', 'trainer', 'human resources', 
            'recruiter', 'office personnel', 'payroll assistant', 'finance', 
            'dispatch', 'management', 'admin', 'survey', 'warehouse', 'walker', 
            'driver', 'helper', 'driver-lite', 'driver-step van', 
            'driver-unscheduled', 'lead driver', 'ddu dedicated', 'ddu shared', 
            'non-dsp related', 'driver -major appliance'
        ]
        
        if pd.isna(val_jt) or str(val_jt).strip() == "":
            missing_fields.append("Job Title")
        elif str(val_jt).strip().lower() not in allowed_titles:
            invalid_jt = True
            
        # Check Work Location
        val_wl = row.get('Work Location')
        if pd.isna(val_wl) or str(val_wl).strip() == "":
            missing_fields.append("Work Location")
            
        if missing_fields or invalid_jt:
            err_reasons = []
            if missing_fields:
                err_reasons.append("Mandatory fields are blank")
            if invalid_jt:
                err_reasons.append(f"Invalid Job Title: '{val_jt}'")
                
            errors.append({
                "Employee ID": emp_id,
                "Missing Fields": ", ".join(missing_fields),
                "Error": " | ".join(err_reasons)
            })
            
    return pd.DataFrame(errors)
    return pd.DataFrame(errors)

def read_uzio_template_df(file):
    """
    Reads the 'Employee Details' sheet from a Uzio template .xlsm.
    Identifies the header row (index 3) and returns the full DataFrame.
    """
    try:
        # Load the whole workbook to preserve everything, but read as DF for logic
        df_template = pd.read_excel(file, sheet_name='Employee Details', header=3, dtype=str)
        # Normalize column names for internal matching (strip space, handles newlines)
        df_template.columns = [str(c).replace("\n", " ").replace("\r", " ").strip() for c in df_template.columns]
        return df_template
    except Exception as e:
        print(f"Error reading Uzio template: {e}")
        return None

def selective_update_uzio(df_source, df_template, selected_uzio_cols, vendor_field_map, fix_options=None):
    """
    Updates specific columns in df_template using data from df_source.
    Only updates employees present in df_source.
    Returns the updated df_template and a summary of changes.
    """
    # 1. Reverse the mapping to find which 'standard field' relates to the selected 'Uzio Column'
    # selected_uzio_cols are the raw keys from UZIO_RAW_MAPPING
    
    # 2. Normalize Employee IDs for matching
    emp_id_col_source = vendor_field_map.get('Employee ID')
    emp_id_col_template = 'Employee ID*' if 'Employee ID*' in df_template.columns else 'Employee ID'
    
    if not emp_id_col_source or emp_id_col_source not in df_source.columns:
        return df_template, "Error: Source 'Employee ID' column not found."
    
    # Prepare lookup dict from source: {id: row_data}
    df_source_clean = df_source.copy()
    df_source_clean[emp_id_col_source] = norm_key_series(df_source_clean[emp_id_col_source])
    # --- FIX: Deduplicate to prevent "DataFrame index must be unique" error ---
    df_source_clean = df_source_clean.drop_duplicates(subset=[emp_id_col_source], keep='first')
    source_lookup = df_source_clean.set_index(emp_id_col_source).to_dict('index')
    
    # 3. Create a temp copy for formatting
    # We use a dummy generator logic to get formatted values for each standard field
    from utils.audit_utils import UZIO_RAW_MAPPING
    
    updated_count = 0
    df_updated = df_template.copy()
    
    # Track changes for display
    change_details = []

    for idx, row in df_updated.iterrows():
        eid = norm_key_series(pd.Series([row.get(emp_id_col_template, "")])).iloc[0]
        
        if eid in source_lookup:
            source_row_data = source_lookup[eid]
            # Create a mini-dataframe for this person to use existing formatters if needed
            # Or just pull logic from generate_uzio_template
            
            for uzio_col in selected_uzio_cols:
                std_name = UZIO_RAW_MAPPING.get(uzio_col)
                if not std_name: continue
                
                vendor_col = vendor_field_map.get(std_name)
                if not vendor_col or vendor_col not in df_source.columns: continue
                
                val = source_row_data.get(vendor_col)
                # Apply same formatting used in generate_uzio_template
                formatted_val = ""
                val_str = str(val).strip().lower() if pd.notna(val) else ""
                if val_str and val_str != "nan":
                    # Reuse specific formatters
                    if std_name == 'Middle Initial':
                        formatted_val = str(val).strip()[0]
                    elif std_name in ['Hire Date', 'Original Hire Date', 'Termination Date', 'DOB']:
                        try:
                            dt = pd.to_datetime(str(val).strip(), errors='coerce')
                            formatted_val = dt.strftime('%d/%m/%Y') if not pd.isna(dt) else str(val).strip()
                        except:
                            formatted_val = str(val).strip()
                    elif std_name == 'License Expiration Date':
                        v_str = str(val).strip()
                        if fix_options and fix_options.get('fix_license', False):
                            if '00/00/0000' in v_str or v_str in ('0', '00', '0000'):
                                formatted_val = ""
                            else:
                                try:
                                    dt = pd.to_datetime(v_str, errors='coerce')
                                    formatted_val = dt.strftime('%m/%d/%Y') if not pd.isna(dt) else ""
                                except:
                                    formatted_val = ""
                        else:
                            formatted_val = v_str
                    elif std_name == 'SSN':
                        formatted_val = str(val).replace("-", "").strip()
                    elif std_name == 'Gender':
                        g_str = str(val).lower()
                        if g_str.startswith('m'): formatted_val = "Male"
                        elif g_str.startswith('f'): formatted_val = "Female"
                    elif std_name == 'Employment Status':
                        s = str(val).lower()
                        if fix_options and fix_options.get('fix_status', False):
                            if 'not hired' in s: formatted_val = 'EXCLUDE'
                            elif 'inactive' in s or 'term' in s: formatted_val = 'TERMINATED'
                            elif 'active' in s or 'leave' in s: formatted_val = 'ACTIVE'
                            else: formatted_val = str(val).strip().upper()
                        else:
                            formatted_val = str(val).strip().upper()
                    elif std_name in ['Zip', 'Mailing Zip']:
                        z_clean = re.sub(r'\D', '', str(val).strip())
                        formatted_val = z_clean.zfill(5)[:5] if z_clean else ""
                    elif std_name == 'Employment Type':
                        et_str = str(val).lower()
                        if fix_options and fix_options.get('fix_type', False):
                            if 'full' in et_str: formatted_val = 'Full Time'
                            elif 'part' in et_str: formatted_val = 'Part Time'
                            elif 'season' in et_str: formatted_val = 'Seasonal'
                            elif 'other' in et_str: formatted_val = 'Other'
                            elif 'intern' in et_str: formatted_val = 'Part Time'
                        else:
                            formatted_val = str(val).strip()
                    elif std_name == 'Termination Reason':
                        tr_str = str(val).strip().lower()
                        if "involuntary" in tr_str or "invluntary" in tr_str:
                            formatted_val = "Involuntary Termination of Employment"
                        elif "voluntary" in tr_str or "quit" in tr_str:
                            formatted_val = "Voluntary Termination of Employment"
                        elif "death" in tr_str:
                            formatted_val = "Death"
                        elif "retire" in tr_str:
                            formatted_val = "Retirement"
                        elif "disability" in tr_str:
                            formatted_val = "Permanent Disability"
                        elif "transfer" in tr_str:
                            formatted_val = "Transfer"
                        else:
                            formatted_val = "Other"
                    else:
                        formatted_val = str(val).strip()
                else:
                    # Value is blank or 'nan'
                    if std_name == 'Job Title':
                        # Check for Job Title fallback
                        if fix_options and (fix_options.get('fix_job_title', False) or fix_options.get('fix_position', False)):
                            dept_col = resolved_field_map.get('Department')
                            if dept_col and dept_col in df_source.columns:
                                dept_val = row.get(dept_col)
                                dept_val_str = str(dept_val).strip().lower() if pd.notna(dept_val) else ""
                                if dept_val_str and dept_val_str != "nan":
                                    formatted_val = str(dept_val).strip()
                                else:
                                    formatted_val = ""
                            else:
                                formatted_val = ""
                        else:
                            formatted_val = ""
                    else:
                        formatted_val = ""
                
                # Check for delete/update
                old_val = row.get(uzio_col, "")
                if str(formatted_val) != str(old_val):
                    df_updated.at[idx, uzio_col] = formatted_val
                    change_details.append({
                        'Employee ID': eid,
                        'Column': uzio_col,
                        'From': old_val,
                        'To': formatted_val
                    })
                    updated_count += 1

            # Special case: Email Fallback (only if Work Email was selected and fix_emails is True)
            if fix_options and fix_options.get('fix_emails', False):
                work_email_col = next((c for c in selected_uzio_cols if UZIO_RAW_MAPPING.get(c) == 'Work Email'), None)
                if work_email_col:
                    current_work_email = df_updated.at[idx, work_email_col]
                    if pd.isna(current_work_email) or str(current_work_email).strip() == "":
                        pers_email_col = next((c for c in df_source.columns if norm_colname(c).casefold() == 'personal email'), None)
                        if not pers_email_col:
                             # Try fuzzy match if direct fails
                             pers_email_col = next((c for c in df_source.columns if 'personal' in c.lower() and 'email' in c.lower()), None)
                        
                        if pers_email_col:
                            fallback_email = str(source_row_data.get(pers_email_col, "")).strip()
                            if fallback_email:
                                df_updated.at[idx, work_email_col] = fallback_email
                                summary_id = f"Email Fallback ({eid})"
                                if not any(d['Employee ID'] == eid and d['Column'] == work_email_col for d in change_details):
                                    change_details.append({
                                        'Employee ID': eid,
                                        'Column': work_email_col,
                                        'From': '(blank)',
                                        'To': fallback_email
                                    })
                                    updated_count += 1

    # --- Post-processing: License Rules on the updated template (Optional) ---
    if fix_options and fix_options.get('fix_license', False):
        lic_num_col = 'License Number*'
        lic_exp_col = 'License Expiration Date'
        if lic_exp_col in df_updated.columns:
            # Clear any 00/00/0000 or invalid placeholders
            bad_mask = df_updated[lic_exp_col].astype(str).str.strip().isin(['00/00/0000', '0', '00', '0000', 'nan', 'NaT', ''])
            df_updated.loc[bad_mask, lic_exp_col] = ""
            # Clear expiration date when license number is blank
            if lic_num_col in df_updated.columns:
                no_lic_mask = df_updated[lic_num_col].isna() | (df_updated[lic_num_col].astype(str).str.strip() == "") | (df_updated[lic_num_col].astype(str).str.strip() == 'nan')
                df_updated.loc[no_lic_mask, lic_exp_col] = ""

    summary = f"Updated {updated_count} employees. Total {len(change_details)} cell changes."
    return df_updated, summary, pd.DataFrame(change_details)

def generate_excel_with_audit(df_main, df_audit, sheet_name_main="Corrected Census", sheet_name_audit="Change Log"):
    """
    Generates an Excel file in memory with two sheets.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_main.to_excel(writer, index=False, sheet_name=sheet_name_main)
        if not df_audit.empty:
            df_audit.to_excel(writer, index=False, sheet_name=sheet_name_audit)
            
            # Formatting the main sheet
            workbook = writer.book
            worksheet_main = writer.sheets[sheet_name_main]
            text_format = workbook.add_format({'num_format': '@'}) # Text format
            
            for col_num, col_name in enumerate(df_main.columns):
                if any(k in str(col_name).lower() for k in ["zip", "postal", "zipcode"]):
                    worksheet_main.set_column(col_num, col_num, 15, text_format)
                else:
                    worksheet_main.set_column(col_num, col_num, 18)

            # Formatting the audit sheet
            worksheet = writer.sheets[sheet_name_audit]
            header_format = workbook.add_format({'bold': True, 'bg_color': '#D7E4BC', 'border': 1})
            for col_num, value in enumerate(df_audit.columns.values):
                worksheet.write(0, col_num, value, header_format)
                if value == "Comments":
                    worksheet.set_column(col_num, col_num, 40) # Wider column for comments
                else:
                    worksheet.set_column(col_num, col_num, 20) # Set column width
        else:
            # If empty, just create an empty sheet with headers
            cols = ["Employee ID", "Employee Name", "Field Changed", "Old Value", "Assumed Value", "Comments"]
            pd.DataFrame(columns=cols).to_excel(writer, index=False, sheet_name=sheet_name_audit)
    
    return output.getvalue()
