import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "ADP vs Uzio – Time Off Tool"

def clean_id(x):
    """Normalize Employee ID (remove .0, strip, remove leading zeros)."""
    if pd.isna(x): return ""
    s = str(x).strip()
    if s.endswith(".0"): s = s[:-2]
    # Remove leading zeros to match typically
    s = s.lstrip("0")
    return s

def run_tool(file_adp, file_uzio):
    # 1. Read ADP Report
    try:
        df_a = pd.read_excel(file_adp)
    except Exception as e:
        st.error(f"Error reading ADP file: {e}")
        return None

    # Normalize ADP Columns
    a_cols = {c.strip().upper(): c for c in df_a.columns}
    
    col_id_a = next((c for c in a_cols.values() if "ASSOCIATE ID" in c.upper()), None)
    col_bal_a = next((c for c in a_cols.values() if "BALANCE AMOUNT" in c.upper()), None)
    col_name_a = next((c for c in a_cols.values() if "NAME" in c.upper() and "POLICY" not in c.upper()), None) # Avoid POLICY NAME

    if not col_id_a or not col_bal_a:
        st.error(f"Could not find required columns in ADP file. Found: {list(df_a.columns)}")
        return None

    # Group by Associate ID and Sum Balance
    # We need to preserve Name if possible, so we'll aggregate Name by taking 'first'
    agg_rules = {col_bal_a: 'sum'}
    if col_name_a:
        agg_rules[col_name_a] = 'first'
        
    # Clean IDs before grouping? Logic:
    # ADP IDs might be consistent, but let's clean them to be safe when mapping
    df_a['Clean_ID'] = df_a[col_id_a].apply(clean_id)
    
    # Filter out empty IDs
    df_a = df_a[df_a['Clean_ID'] != ""]
    
    # Check if 'Clean_ID' exists before groupby
    if df_a.empty:
        st.error("No valid Employee IDs found in ADP file.")
        return None

    df_grouped = df_a.groupby('Clean_ID').agg(agg_rules).reset_index()
    
    # Create Lookup Map: CleanID -> Total Balance
    balance_map = {}
    name_map = {}
    
    for idx, row in df_grouped.iterrows():
        eid = row['Clean_ID']
        val = row[col_bal_a]
        name = row[col_name_a] if col_name_a else "N/A"
        
        balance_map[eid] = val
        name_map[eid] = name

    # ---------------------------------------------------------
    # PART A: Generate Clean Import File (using openpyxl)
    # ---------------------------------------------------------
    file_uzio.seek(0)
    try:
        wb = openpyxl.load_workbook(file_uzio)
    except Exception as e:
        st.error(f"Error reading Uzio Template with openpyxl: {e}")
        return None

    if len(wb.sheetnames) < 2:
        st.error("Uzio Template must have at least 2 sheets (Instruction, Time Off Details).")
        return None
        
    ws = wb.worksheets[1] # "Time Off Details"
    
    header_row = 4
    idx_id_u = None
    idx_bal_u = None
    
    for cell in ws[header_row]:
        val = str(cell.value).strip() if cell.value else ""
        if "Employee ID" in val:
            idx_id_u = cell.column
        elif "Opening Balance" in val: # User confirmed Opening Balance
            idx_bal_u = cell.column
            
    if not idx_id_u or not idx_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening Balance' headers in Row 4 of Sheet 2.")
        return None

    # Iterate Data Rows
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=idx_id_u)
        cell_bal = ws.cell(row=row_idx, column=idx_bal_u)
        
        current_val = cell_bal.value
        
        # Rule: If Blank -> Keep Blank
        if current_val is None or str(current_val).strip() == "":
            continue # Skip
            
        # Policy Assigned -> Update
        eid = clean_id(cell_id.value)
        if eid in balance_map:
            cell_bal.value = balance_map[eid]
            
    # ---------------------------------------------------------
    # PART B: Calculate Audit Data (using pandas)
    # ---------------------------------------------------------
    file_uzio.seek(0)
    try:
        df_u = pd.read_excel(file_uzio, sheet_name=1, header=3)
    except Exception as e:
        st.error(f"Error reading Uzio Template for audit: {e}")
        return None
        
    u_cols = {c.strip(): c for c in df_u.columns}
    col_id_u = next((c for c in u_cols if "Employee ID" in c), None)
    col_bal_u = next((c for c in u_cols if "Opening Balance" in c), None) # Check Opening Balance first
    if not col_bal_u:
         col_bal_u = next((c for c in u_cols if "Operating Balance" in c), None)
         
    col_name_u = next((c for c in u_cols if "Employee Name" in c or "Name" in c), None)

    if not col_id_u or not col_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening Balance' in Uzio Template for audit. Found: {list(df_u.columns)}")
        return None

    # Trackers
    matched_adp_ids = set()
    unassigned_policies_rows = [] 
    all_exceptions = []

    def audit_scan(row):
        current_val = row[col_bal_u]
        if pd.isna(current_val) or str(current_val).strip() == "":
            unassigned_policies_rows.append(row.to_dict())
            
            # Add to Exception
            eid = str(row[col_id_u]) if pd.notna(row[col_id_u]) else ""
            name = str(row[col_name_u]) if col_name_u and pd.notna(row[col_name_u]) else "N/A"
            all_exceptions.append({
                'Employee ID': eid,
                'Employee Name': name,
                'Issue Category': 'Unassigned Policy (Blank Balance)',
                'ADP Balance': ''
            })
            return
            
        eid = clean_id(row[col_id_u])
        if eid in balance_map:
            matched_adp_ids.add(eid)

    df_u.apply(audit_scan, axis=1)

    # Missing in Uzio
    missing_in_uzio = []
    
    # Iterate through the grouped ADP data
    for idx, row in df_grouped.iterrows():
        eid = row['Clean_ID']
        val = row[col_bal_a]
        name = row[col_name_a] if col_name_a else "N/A"

        if eid and eid not in matched_adp_ids:
            if pd.notna(val) and str(val).strip() != "":
                # Reconstruct a row-like dict for the report
                missing_row = {
                    'Employee ID': eid,
                    'Employee Name': name,
                    'Total Balance': val
                }
                missing_in_uzio.append(missing_row)
                
                # Add to Exception
                all_exceptions.append({
                    'Employee ID': eid,
                    'Employee Name': name,
                    'Issue Category': 'Missing in Uzio Template',
                    'ADP Balance': val
                })
    
    df_missing = pd.DataFrame(missing_in_uzio)
    if df_missing.empty: df_missing = pd.DataFrame({'Message': ['All ADP employees matched']})

    df_unassigned = pd.DataFrame(unassigned_policies_rows)
    if df_unassigned.empty: df_unassigned = pd.DataFrame({'Message': ['No unassigned policies found']})

    df_exceptions = pd.DataFrame(all_exceptions)
    if df_exceptions.empty: df_exceptions = pd.DataFrame({'Message': ['No exceptions found']})

    # ---------------------------------------------------------
    # PART C: Append Audit Sheets
    # ---------------------------------------------------------
    def add_sheet_with_df(workbook, sheet_name, dataframe):
        ws_new = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws_new.append(r)

    add_sheet_with_df(wb, "Missing in Uzio", df_missing)
    add_sheet_with_df(wb, "Unassigned Policies", df_unassigned)
    add_sheet_with_df(wb, "ADP Grouped Data", df_grouped)
    add_sheet_with_df(wb, "Exception Summary", df_exceptions)

    out_final = io.BytesIO()
    wb.save(out_final)
    
    return out_final.getvalue()

def render_ui():
    st.title(APP_TITLE)
    client_name = st.text_input("Client Name", value="Client", key="adp_timeoff_client")

    st.markdown("""
    **Instructions**:
    1. Upload **ADP Time Off Balance Summary** (.xlsx).
    2. Upload **Uzio Time Off Import Template** (.xlsx).
    
    **Output**:
    - Generates a **Consolidated Excel File**.
    - **Original Tabs** (Instructions, Time Off Details) are preserved.
    - **Time Off Details** is updated with ADP balances (Summed by Associate ID).
    - **New Tabs** added for Audit: Missing, Unassigned, Raw Data.
    - **Exception Summary** tab at the end.
    """)

    col1, col2 = st.columns(2)
    with col1:
        f_a = st.file_uploader("ADP Balance Summary", type=["xlsx"], key="at_a")
    with col2:
        f_u = st.file_uploader("Uzio Template", type=["xlsx"], key="at_u")

    if st.button("Generate Consolidated Report", key="run_timeoff_adp"):
        if not f_u or not f_a:
            st.error("Please upload both files.")
            return
            
        try:
            with st.spinner("Processing..."):
                res = run_tool(f_a, f_u)
                
            if res:
                timestamp = pd.Timestamp.now().strftime('%d_%m_%Y_%H%M')
                filename = f"{client_name}_Uzio_ADP_TimeOff_Audit_Report_{timestamp}.xlsx"

                st.success("File Generated Successfully!")
                st.download_button(
                    "Download Consolidated File",
                    data=res,
                    file_name=filename
                )
            else:
                st.error("No file could be generated due to an error.")

        except Exception as e:
            st.error(f"An error occurred: {e}")
            st.exception(e)
