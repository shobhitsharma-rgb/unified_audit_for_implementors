"""ADP - Prior Payroll Sanity Check Tool.

Cleans an ADP Prior Payroll file before it's ingested by downstream APIs.
Three independent cleanups are applied as needed:

  1. Per-pay-period aggregation: when the file is exported with one row per
     pay period per employee (the most common implementor mistake), all
     rows for an Associate ID are aggregated into a single row -- money
     and hours columns are SUMmed, period dates are MIN/MAX'd, identity
     columns kept as-is.
  2. Duplicate row merge: when two rows share the same Employee + Pay Date
     (typically a skeleton + detail pair within one pay period), they are
     smart-merged into one row without double-counting.
  3. Grand-total row removal: the last row of the file, where the last
     employee's ID got bled into the totals row, is detected and dropped.
  4. 401k / Roth memo split: when the file has 401k and/or Roth deferral
     columns (LOAN columns never count), the MEMO column carrying the
     combined employer match is identified (its entry count equals the
     number of employees having any 401k or Roth deferral; 0.00 counts as
     an entry) and split per row by the employee's own deferrals:
     401k-only -> match stays put; Roth-only -> match moves entirely to a
     new 'ROTH:<memo column>' column (UPPERCASE prefix -- required by the
     UZIO import's case-sensitive mapping lookup); both -> match stays up
     to the summed 401k deferral (e.g. K-ADP 401K + 28-ADP 401K FLAT$),
     only the excess moves to Roth; neither -> match stays and the row is
     flagged for review. Ties and count mismatches are flagged so the
     user picks the memo column (or skips) before running.
  5. Pay-period date columns: consolidated quarter files often arrive
     without PERIOD BEGINNING DATE / PERIOD ENDING DATE / PAY DATE, which
     the downstream API requires. When any of the three is missing, the
     dates are read from the filename (three 8-digit MMDDYYYY blocks
     joined by underscores, in begin_end_pay order, e.g.
     PriorPayroll_01012026_01072026_01142026.xlsx) and the missing
     columns are inserted between WORKED IN STATE and GROSS PAY, stamped
     on every row as MM/DD/YYYY text. If the filename has no parseable
     dates, the user enters the three dates manually in the UI. Columns
     already present are never modified.
  6. Lived-in State / Local tax split: ADP lumps out-of-jurisdiction
     residents' withholding into single LIVED-IN STATE / LIVED-IN LOCAL
     - EMPLOYEE TAX columns, which cannot be mapped to multiple taxes
     downstream. With the optional Tax Validation Report uploaded, each
     employee's amount moves into a 'LIVED-IN STATE (<abbr>) - EMPLOYEE TAX'
     column (state abbreviation / local jurisdiction description) per Associate
     ID -- only jurisdictions carrying money get a column. The base column
     is deleted once it holds no money (kept only while unmatched values
     remain in it); TAXABLE stays combined; unmatched employees keep their
     value in the base column and are flagged.

Optional Carvan-specific NET PAY <-> TAKE HOME value swap is exposed as
a checkbox in the UI (default ON) because the API expects them reversed.

Output is always CSV with the input's exact column headers and column
order preserved. Input accepts .xlsx / .xls / .csv.
"""

import re
import io
from datetime import datetime

import streamlit as st
import pandas as pd
import openpyxl
from utils.audit_utils import clean_money_val


def _find_col(df, candidates):
    """Case-insensitive exact-then-substring lookup of a column."""
    for cand in candidates:
        for c in df.columns:
            if str(c).strip().lower() == cand.lower():
                return c
    for cand in candidates:
        for c in df.columns:
            if cand.lower() in str(c).strip().lower():
                return c
    return None


# Canonical headers for the API-required date columns, in insertion order.
PERIOD_DATE_COLUMNS = ["PERIOD BEGINNING DATE", "PERIOD ENDING DATE", "PAY DATE"]

_FILENAME_DATES_RE = re.compile(r"(?<!\d)(\d{8})_(\d{8})_(\d{8})(?!\d)")


def check_period_date_columns(df):
    """Report which of the three API-required date columns exist in the file.

    Matching reuses _find_col semantics (case-insensitive, substring), so an
    existing 'Pay Date' or 'Check Date' counts as PAY DATE being present.
    Returns (found_map {canonical: actual_col_or_None}, missing_list).
    """
    found = {
        "PERIOD BEGINNING DATE": _find_col(df, ["Period Beginning Date", "Period Begin Date"]),
        "PERIOD ENDING DATE": _find_col(df, ["Period Ending Date", "Period End Date"]),
        "PAY DATE": _find_col(df, ["Pay Date", "Check Date"]),
    }
    missing = [c for c in PERIOD_DATE_COLUMNS if found[c] is None]
    return found, missing


def parse_filename_dates(filename):
    """Extract the three pay-period dates from the uploaded file's name.

    The name must contain exactly one run of three 8-digit MMDDYYYY blocks
    joined by underscores, in <begin>_<end>_<pay> order, e.g.
    'PriorPayroll_01012026_01072026_01142026.xlsx'. Each block must be a
    real calendar date. Returns {canonical_col: 'MM/DD/YYYY'} or None.
    """
    matches = _FILENAME_DATES_RE.findall(str(filename or ""))
    if len(matches) != 1:
        return None
    out = {}
    for col, block in zip(PERIOD_DATE_COLUMNS, matches[0]):
        try:
            out[col] = datetime.strptime(block, "%m%d%Y").strftime("%m/%d/%Y")
        except ValueError:
            return None
    return out


def insert_period_date_columns(df, dates, missing):
    """Insert the missing date columns between WORKED IN STATE and GROSS PAY,
    stamping every row with the same MM/DD/YYYY string. Columns already in the
    file are never touched. Falls back to inserting before GROSS PAY, then to
    after WORKED IN STATE, then to appending at the end.

    Returns (new_df, {"added": [cols], "placement": key}).
    """
    df = df.copy()
    gross_col = _find_col(df, ["Gross Pay"])
    state_col = _find_col(df, ["Worked In State"])
    if gross_col is not None:
        pos = df.columns.get_loc(gross_col)
        placement = "between" if state_col is not None else "before_gross"
    elif state_col is not None:
        pos = df.columns.get_loc(state_col) + 1
        placement = "after_state"
    else:
        pos = len(df.columns)
        placement = "appended"

    added = []
    for col in PERIOD_DATE_COLUMNS:
        if col in missing:
            df.insert(pos, col, dates[col])
            pos += 1
            added.append(col)
    return df, {"added": added, "placement": placement}


# ─────────────────────────────────────────────────────────────────────────────
# Lived-in State / Local tax split (optional Tax Validation Report upload)
#
# ADP lumps every out-of-state/local resident's withholding into single
# LIVED-IN STATE / LIVED-IN LOCAL - EMPLOYEE TAX columns. One column cannot be
# mapped to multiple taxes downstream, so when a Tax Validation Report is
# uploaded the tool splits the column: one new column per jurisdiction that
# actually carries money — named 'LIVED-IN STATE (<code>) - EMPLOYEE TAX'
# (state abbreviation for states, Lived-in Local Jurisdiction Description for
# locals; the code sits BEFORE '- EMPLOYEE TAX' so downstream tools still
# classify the column as a tax) — inserted right
# after the base column, values moved per Associate ID. The base column is
# deleted once it holds no money (it survives only while unmatched values
# remain in it); the TAXABLE column is intentionally NOT split. Employees
# with money but no jurisdiction in the report keep their value in the base
# column and are flagged for review.
# ─────────────────────────────────────────────────────────────────────────────

# kind -> (ADP base column, validation-report lookup column)
LIVED_IN_SPLIT_CONFIG = {
    "state": ("LIVED-IN STATE - EMPLOYEE TAX", "Lived in State Code"),
    "local": ("LIVED-IN LOCAL - EMPLOYEE TAX", "Lived in Local Jurisdiction Description"),
}


def _norm_header(s):
    """Lowercase and collapse all non-alphanumerics so 'LIVED-IN STATE' and
    'Lived in State' compare equal — and TAX never matches TAXABLE."""
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()).strip()


def _find_exact_col(df, target):
    """Exact (normalized) header match only — no substring fallback, so the
    LIVED-IN ... TAX column can never accidentally resolve to ... TAXABLE."""
    want = _norm_header(target)
    for c in df.columns:
        if _norm_header(c) == want:
            return c
    return None


def find_lived_in_columns(df):
    """{kind: actual_column_name} for the lived-in tax columns present."""
    out = {}
    for kind, (base, _lookup) in LIVED_IN_SPLIT_CONFIG.items():
        col = _find_exact_col(df, base)
        if col is not None:
            out[kind] = col
    return out


def _is_movable_value(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    return str(v).strip() not in ("", "-", "nan", "NaT", "None")


def load_tax_validation_report(file):
    """Parse the ADP Tax Validation Report (.csv/.xlsx) into per-kind lookups:
    {'state': {associate_id: 'WI', ...}, 'local': {...}}.

    Duplicate Associate IDs are common (re-hires); the first NON-BLANK value
    per ID wins. Returns (lookups, error_message_or_None)."""
    file.seek(0)
    name = (getattr(file, "name", "") or "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(file, dtype=str)
    else:
        df = pd.read_excel(file, dtype=str)

    eid_col = _find_exact_col(df, "Associate ID")
    if eid_col is None:
        return None, "No 'Associate ID' column found in the Tax Validation Report."

    lookups = {}
    for kind, (_base, lookup_col) in LIVED_IN_SPLIT_CONFIG.items():
        col = _find_exact_col(df, lookup_col)
        table = {}
        if col is not None:
            for _, row in df.iterrows():
                eid = str(row[eid_col]).strip()
                val = row[col]
                if not eid or eid.lower() == "nan":
                    continue
                if eid in table:
                    continue  # first non-blank value wins
                if _is_movable_value(val):
                    table[eid] = str(val).strip()
        lookups[kind] = table
    return lookups, None


_LIVED_IN_TAX_SUFFIX_RE = re.compile(r"^(?P<name>.*?)(?P<suffix>\s*-\s*EMPLOY(?:EE|ER)\s+TAX)\s*$",
                                     re.IGNORECASE)


def _jurisdiction_col_name(base_col, code):
    """'LIVED-IN STATE - EMPLOYEE TAX' + 'IL' -> 'LIVED-IN STATE (IL) - EMPLOYEE TAX'.

    The '(code)' goes BEFORE the '- EMPLOYEE TAX' suffix so the new column still
    ends with 'EMPLOYEE TAX' — the shape every downstream tool (Setup Helper
    column classifier, audit unmapped-column scan) recognizes as a tax column."""
    m = _LIVED_IN_TAX_SUFFIX_RE.match(str(base_col).strip())
    if m:
        return f"{m.group('name')} ({code}){m.group('suffix')}"
    return f"{base_col} ({code})"


def split_lived_in_column(df, base_col, lookup):
    """Move each row's lived-in tax value into a per-jurisdiction column
    ('LIVED-IN STATE (IL) - EMPLOYEE TAX') per the Associate ID -> jurisdiction
    lookup. Only jurisdictions that actually receive money get a column
    (inserted right after base_col, sorted).
    Unmatched employees keep their value in the base column.

    Once the base column holds no money (everything moved, or it was empty to
    begin with) it is DELETED — it is kept only while unmatched values remain.

    Returns (new_df, {"created": [cols], "moved": n, "dropped_base": bool,
                      "unmatched": [{"Associate ID", "Amount"}]})."""
    df = df.copy()
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    if eid_col is None:
        return df, {"created": [], "moved": 0, "dropped_base": False, "unmatched": []}

    moves = []       # (row_idx, code, value)
    unmatched = []
    for idx in df.index:
        v = df.at[idx, base_col]
        if not _is_movable_value(v):
            continue
        eid = str(df.at[idx, eid_col]).strip()
        code = lookup.get(eid, "")
        if code:
            moves.append((idx, code, v))
        else:
            unmatched.append({"Associate ID": eid, "Amount": v})

    created = []
    if moves:
        codes = sorted({code for _, code, _ in moves})
        pos = df.columns.get_loc(base_col) + 1
        for code in codes:
            new_col = _jurisdiction_col_name(base_col, code)
            if new_col not in df.columns:
                df.insert(pos, new_col, "")
                created.append(new_col)
                pos += 1
            else:
                pos = df.columns.get_loc(new_col) + 1
        for idx, code, v in moves:
            df.at[idx, _jurisdiction_col_name(base_col, code)] = v
            df.at[idx, base_col] = ""

    dropped_base = False
    if not df[base_col].map(_is_movable_value).any():
        df = df.drop(columns=[base_col])
        dropped_base = True

    return df, {"created": created, "moved": len(moves),
                "dropped_base": dropped_base, "unmatched": unmatched}


_ROUND_FORMULA_RE = re.compile(r"^=ROUND\(\s*(-?[\d.]+)\s*,\s*([\d.]+)\s*\)\s*$", re.IGNORECASE)


def _evaluate_cell(value):
    """Resolve =ROUND(x,n) formulas (the only formula style ADP exports use for money cells).

    Returns the ROUNDED numeric value when the cell holds such a formula --
    honoring the formula's digit count, exactly as Excel would display it
    (=ROUND(15.669817, 2.0) -> 15.67, not 15.669817). Otherwise returns the
    cell value unchanged. Pandas read_excel sees these formula cells as None,
    so we read with openpyxl and feed each cell through this evaluator.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s.startswith("="):
        return value
    m = _ROUND_FORMULA_RE.match(s)
    if m:
        try:
            digits = int(float(m.group(2)))
            return round(float(m.group(1)), digits)
        except ValueError:
            return None
    return None


def _read_excel_with_formula_eval(file):
    """Read .xlsx/.xls via openpyxl, evaluating =ROUND() formulas to their literal values.

    Returns (df, header_idx, sheet_name). Picks the first non-criteria sheet, finds
    the header row by searching for an ID column header, and parses every data cell.
    """
    file.seek(0)
    wb = openpyxl.load_workbook(file, data_only=False)
    target_sheet = wb.sheetnames[0]
    if len(wb.sheetnames) > 1 and "criteria" in wb.sheetnames[0].lower():
        target_sheet = wb.sheetnames[1]
    ws = wb[target_sheet]

    header_idx = 0
    for r in range(1, min(ws.max_row, 50) + 1):
        row_text = " ".join(
            str(ws.cell(r, c).value).lower()
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        )
        if any(k in row_text for k in ["associate id", "employee id", "file #"]):
            header_idx = r - 1
            break

    headers = [ws.cell(header_idx + 1, c).value for c in range(1, ws.max_column + 1)]
    headers = _dedup_headers(headers)
    rows = []
    for r in range(header_idx + 2, ws.max_row + 1):
        row = [_evaluate_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        rows.append(row)
    df = pd.DataFrame(rows, columns=headers)
    return df, header_idx, target_sheet


def _dedup_headers(headers):
    """Match pandas.read_csv's mangle-dupe-cols behavior: first occurrence keeps
    the original name, subsequent occurrences get '.1', '.2', etc. suffixes.
    The rest of the pipeline (aggregate_by_associate, etc.) already expects this.
    """
    seen = {}
    out = []
    for h in headers:
        h_str = "" if h is None else str(h)
        if h_str in seen:
            seen[h_str] += 1
            out.append(f"{h_str}.{seen[h_str]}")
        else:
            seen[h_str] = 0
            out.append(h_str)
    return out


_EXCESS_DECIMALS_RE = re.compile(r"^-?\d+\.\d{3,}$")


def normalize_money_precision(df):
    """Normalize float noise to 2 decimals for the output CSV.

    Excel stores numbers as binary doubles (769.34 is really
    769.33999999999997) and older cleaned files carry full-precision tax
    values -- the CSV should show money the way Excel displays it. Floats
    round to 2. Strings are touched ONLY when they are a plain decimal number
    with 3+ decimal places (IDs, dates, zips, SSNs never match that shape).
    Ints and everything else pass through unchanged."""
    def fix(v):
        if isinstance(v, float):
            return v if pd.isna(v) else round(v, 2)
        if isinstance(v, str) and _EXCESS_DECIMALS_RE.match(v.strip()):
            try:
                return f"{float(v):.2f}"
            except ValueError:
                return v
        return v
    return df.apply(lambda col: col.map(fix))


def read_input_file(file):
    """Read the ADP file (xlsx/xls/csv), find the header row, and return the dataframe.

    Preserves original column names and order exactly. Does NOT strip the grand-total
    row or summary rows -- the sanity-check pipeline does that explicitly so each
    cleanup step can be reported.
    """
    file.seek(0)
    name = (file.name or "").lower()

    if name.endswith(".csv"):
        file.seek(0)
        df_peek = pd.read_csv(file, header=None, nrows=50, dtype=str)
        header_idx = 0
        for i, row in df_peek.iterrows():
            row_str = " ".join(str(x).lower() for x in row if pd.notna(x))
            if any(k in row_str for k in ["associate id", "employee id", "file #"]):
                header_idx = i
                break
        file.seek(0)
        df = pd.read_csv(file, header=header_idx, dtype=str)
        return df, header_idx, "Sheet1"

    return _read_excel_with_formula_eval(file)


def drop_summary_rows(df):
    """Drop the per-employee 'Totals For Associate ID XYZ:' rows the ADP report
    interleaves between pay-period rows. They have a null Associate ID and all
    money columns blank, so they're useless once we re-aggregate ourselves.

    Returns (cleaned_df, removed_count).
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    if not eid_col:
        return df.reset_index(drop=True), 0
    mask = df[eid_col].notna() & (df[eid_col].astype(str).str.strip() != "")
    removed = (~mask).sum()
    return df[mask].reset_index(drop=True), int(removed)


def detect_per_pay_period_structure(df):
    """Decide whether the file needs per-associate aggregation.

      - 'aggregate' : at least one associate has more than one row -- this is the
        per-pay-period export the implementor often produces. Roll up to one row
        per associate (sums money/hours, min/max for dates, identity columns kept).
        Same-pay-date duplicates -- common in ADP when an employee gets two checks
        on the same day, each row carrying real values -- are also folded together
        by the SUM aggregation, which is the correct behavior for ADP.
      - 'none'      : already clean -- one row per associate.

    Returns (mode, summary_dict).
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date"])
    if not eid_col:
        return "none", None

    work = df[df[eid_col].notna()].copy()
    work[eid_col] = work[eid_col].astype(str).str.strip()
    work = work[work[eid_col] != ""]
    if work.empty:
        return "none", None

    rows_per_eid = work.groupby(eid_col).size()
    total_associates = int(len(rows_per_eid))
    multi_row = int((rows_per_eid > 1).sum())

    summary = {
        "associates": total_associates,
        "with_multiple_rows": multi_row,
        "max_rows_for_single_associate": int(rows_per_eid.max()),
    }
    if pay_col:
        pay_dates_per_eid = work.groupby(eid_col)[pay_col].nunique()
        summary["max_pay_dates_for_single_associate"] = int(pay_dates_per_eid.max())

    return ("aggregate" if multi_row > 0 else "none"), summary


def _to_float(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "nan", "NaT"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _format_date(dt):
    if pd.isna(dt):
        return None
    try:
        return dt.strftime("%m/%d/%Y")
    except Exception:
        return None


def aggregate_by_associate(df):
    """Collapse a per-pay-period file into one row per Associate ID.

    Aggregation rules:
      - Money / hours / earning / tax columns: SUM
      - PERIOD BEGINNING DATE: MIN
      - PERIOD ENDING DATE / PAY DATE / TERMINATION DATE: MAX
      - Identity columns (NAME, FILE NUMBER, POSITION ID, STATUS, TAX ID,
        DIST #, WORKED IN STATE): first non-null value
      - CHECK/VOUCHER NUMBER: blanked (pay-period specific)
      - Anything else: SUM if numeric, else first non-null

    Returns (aggregated_df, summary_dict).
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date"])
    period_begin_col = _find_col(df, ["Period Beginning Date", "Period Begin Date", "Start Date"])
    period_end_col = _find_col(df, ["Period Ending Date", "Period End Date", "End Date"])
    term_col = _find_col(df, ["Termination Date"])
    check_col = _find_col(df, ["Check/Voucher Number", "Check Number", "Voucher Number"])

    min_date_cols = {period_begin_col} - {None}
    max_date_cols = {period_end_col, pay_col, term_col} - {None}

    # Identity columns are constant per employee -- never summed even though some
    # (FILE NUMBER, DIST #) parse as numeric strings. Use the first non-null value.
    identity_col_names = ["Name", "File Number", "Position ID", "Status", "Tax ID",
                          "Dist #", "Worked In State"]
    identity_cols = {_find_col(df, [n]) for n in identity_col_names} - {None}

    if not eid_col:
        return df, None

    EMPTY_PLACEHOLDER = "-"

    aggregated_rows = []
    for eid_val, group in df.groupby(eid_col, sort=False):
        out_row = {}
        for col in df.columns:
            vals = group[col].tolist()

            if col == eid_col:
                out_row[col] = eid_val
                continue
            if col == check_col:
                out_row[col] = ""
                continue

            if col in min_date_cols or col in max_date_cols:
                dts = pd.to_datetime(vals, errors="coerce")
                dts = dts.dropna() if hasattr(dts, "dropna") else pd.Series(dts).dropna()
                if len(dts) == 0:
                    out_row[col] = EMPTY_PLACEHOLDER
                else:
                    target = dts.min() if col in min_date_cols else dts.max()
                    out_row[col] = _format_date(target) or EMPTY_PLACEHOLDER
                continue

            if col in identity_cols:
                # Take the first non-null value (constant per employee)
                first = next(
                    (v for v in vals
                     if v is not None
                     and not (isinstance(v, float) and pd.isna(v))
                     and str(v).strip() not in ("", "nan", "NaT")),
                    None,
                )
                out_row[col] = first if first is not None else EMPTY_PLACEHOLDER
                continue

            numeric_vals = []
            categorical_vals = []
            for v in vals:
                f = _to_float(v)
                if f is not None:
                    numeric_vals.append(f)
                elif v is not None and not (isinstance(v, float) and pd.isna(v)):
                    s = str(v).strip()
                    if s and s.lower() not in ("nan", "nat"):
                        categorical_vals.append(v)

            if numeric_vals and not categorical_vals:
                if all(v == 0 for v in numeric_vals):
                    out_row[col] = EMPTY_PLACEHOLDER
                else:
                    out_row[col] = round(sum(numeric_vals), 2)
            elif categorical_vals:
                out_row[col] = categorical_vals[0]
            else:
                out_row[col] = EMPTY_PLACEHOLDER

        aggregated_rows.append(out_row)

    out_df = pd.DataFrame(aggregated_rows, columns=df.columns)
    return out_df, {
        "input_rows": int(len(df)),
        "output_rows": int(len(out_df)),
        "associates": int(len(out_df)),
    }


def apply_net_take_swap(df):
    """Swap NET PAY <-> TAKE HOME column values. The Carvan-style API maps these
    reversed, so when the swap is enabled the data ends up under the API's expected
    semantics. Column headers are NOT changed -- only the values are exchanged.
    """
    net_col = _find_col(df, ["Net Pay"])
    take_col = _find_col(df, ["Take Home"])
    if not net_col or not take_col or net_col == take_col:
        return df, False
    net_vals = df[net_col].copy()
    df[net_col] = df[take_col].copy()
    df[take_col] = net_vals
    return df, True


def detect_grand_total_row(df):
    """Detect the bottom-of-file grand total where the last employee's ID leaked.

    Pattern (carried over from the audit tool):
      - last row's first few columns share values with the previous row
        (the leak), AND
      - some money column on the last row equals the sum of all preceding
        rows for that column (within 5%).

    Returns (cleaned_df, info_dict_or_None).
    """
    if len(df) < 2:
        return df, None

    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]

    shared = 0
    for c in df.columns[:5]:
        v_l = str(last_row[c]).strip()
        v_p = str(prev_row[c]).strip()
        if v_l and v_l == v_p and v_l.lower() != "nan":
            shared += 1
    if shared < 1:
        return df, None

    for c in df.columns:
        try:
            val_last = clean_money_val(last_row[c])
            if val_last <= 100:
                continue
            sum_rest = sum(clean_money_val(x) for x in df[c].iloc[:-1])
            if sum_rest > 0 and abs(val_last - sum_rest) < sum_rest * 0.05:
                eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
                first_col = _find_col(df, ["First Name"])
                last_col = _find_col(df, ["Last Name"])
                preview_eid = str(last_row[eid_col]) if eid_col else ""
                fn = str(last_row[first_col]).strip() if first_col and pd.notna(last_row[first_col]) else ""
                ln = str(last_row[last_col]).strip() if last_col and pd.notna(last_row[last_col]) else ""
                return df.iloc[:-1].copy(), {
                    "removed_employee_id": preview_eid,
                    "removed_employee_name": (fn + " " + ln).strip(),
                    "matched_on_column": str(c),
                    "matched_value": round(val_last, 2),
                    "expected_sum": round(sum_rest, 2),
                }
        except Exception:
            continue

    return df, None


def _smart_merge_value(values):
    """Pick the best value across duplicate rows for a single column.

    Rules:
      - Drop NaN / empty / dash placeholders
      - Among numeric candidates, take the one with the largest absolute value
        (avoids double-counting when one row is the skeleton 0 / dash row)
      - For non-numeric, take the first non-empty value
      - Fall back to the first raw value if everything is empty
    """
    cleaned = []
    for v in values:
        if pd.isna(v):
            continue
        sv = str(v).strip()
        if sv in ("", "-", "nan", "NaT"):
            continue
        cleaned.append(v)
    if not cleaned:
        return values[0] if len(values) > 0 else None

    best_num = None
    best_num_val = None
    for v in cleaned:
        try:
            num = clean_money_val(v)
            if best_num is None or abs(num) > abs(best_num_val):
                best_num = v
                best_num_val = num
        except Exception:
            continue
    if best_num is not None and best_num_val is not None and best_num_val != 0:
        return best_num
    return cleaned[0]


def merge_duplicate_pay_periods(df):
    """Fold duplicate (Employee ID, Pay Date [, Period Start, Period End]) rows
    into one row using smart merge.

    Returns (cleaned_df, list_of_merge_events).
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date", "Pay Period End Date"])
    start_col = _find_col(df, ["Period Start", "Pay Period Start", "Start Date"])
    end_col = _find_col(df, ["Period End", "Pay Period End", "End Date"])

    if not eid_col or not pay_col:
        return df, []

    keys = [eid_col, pay_col]
    if start_col and start_col not in keys:
        keys.append(start_col)
    if end_col and end_col not in keys:
        keys.append(end_col)

    work = df.copy()
    work["_orig_idx"] = range(len(work))

    grouped = work.groupby(keys, dropna=False, sort=False)
    counts = grouped.size().reset_index(name="_n")
    dup_keys = counts[counts["_n"] > 1]
    if dup_keys.empty:
        return df.reset_index(drop=True), []

    keep_indices = []
    drop_indices = set()
    merge_events = []
    merged_records = []

    for key_vals, group in grouped:
        if len(group) == 1:
            keep_indices.append(group["_orig_idx"].iloc[0])
            continue

        first_idx = int(group["_orig_idx"].iloc[0])
        merged = {}
        for col in df.columns:
            merged[col] = _smart_merge_value(group[col].tolist())

        merged_records.append((first_idx, merged))
        drop_indices.update(int(i) for i in group["_orig_idx"].tolist())

        merge_events.append({
            "Employee ID": str(key_vals[0]),
            "Pay Date": str(key_vals[1]),
            "Rows merged": int(len(group)),
            "Kept canonical row at original index": first_idx,
        })

    cleaned_rows = []
    for i in range(len(df)):
        if i in drop_indices:
            continue
        cleaned_rows.append(df.iloc[i].to_dict())
    for first_idx, merged in merged_records:
        merged["_insert_at"] = first_idx
        cleaned_rows.append(merged)

    cleaned_rows.sort(key=lambda r: r.get("_insert_at", -1) if "_insert_at" in r else -1)
    for r in cleaned_rows:
        r.pop("_insert_at", None)

    cleaned = pd.DataFrame(cleaned_rows, columns=df.columns)
    return cleaned.reset_index(drop=True), merge_events


def detect_file_shape(df):
    """Inspect a (cleaned) ADP Prior Payroll DataFrame and return facts +
    a recommended aggregation_strategy. Read-only: never mutates df. Mirrors
    the audit_fast_api version of the same name; keep both in sync.
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date"])
    pbeg_col = _find_col(df, ["Period Beginning Date", "Period Begin Date", "Start Date"])
    pend_col = _find_col(df, ["Period Ending Date", "Period End Date", "End Date"])

    facts = {
        "associates": 0, "total_rows": int(len(df)),
        "rows_per_associate_max": 0, "rows_per_associate_avg": 0.0,
        "distinct_pay_dates": 0, "distinct_pay_dates_per_associate_avg": 0.0,
        "period_min": None, "period_max": None, "date_span_days": None,
        "detected_shape": "unknown",
        "recommended_strategy": None,
        "recommendation_reason": "",
    }
    if not eid_col:
        facts["recommendation_reason"] = "No Associate ID column found; cannot recommend a strategy."
        return facts

    work = df[df[eid_col].notna()].copy()
    work[eid_col] = work[eid_col].astype(str).str.strip()
    work = work[work[eid_col] != ""]
    if work.empty:
        facts["recommendation_reason"] = "No data rows with an Associate ID."
        return facts

    rows_per_eid = work.groupby(eid_col).size()
    facts["associates"] = int(len(rows_per_eid))
    facts["rows_per_associate_max"] = int(rows_per_eid.max())
    facts["rows_per_associate_avg"] = round(float(rows_per_eid.mean()), 2)

    if pay_col:
        pay_parsed = pd.to_datetime(work[pay_col], errors="coerce")
        facts["distinct_pay_dates"] = int(pay_parsed.dropna().nunique())
        per_eid_pd = work.assign(_pd=pay_parsed).groupby(eid_col)["_pd"].nunique()
        facts["distinct_pay_dates_per_associate_avg"] = round(float(per_eid_pd.mean()), 2)

    pmin = pmax = None
    if pbeg_col:
        s = pd.to_datetime(work[pbeg_col], errors="coerce").dropna()
        if not s.empty: pmin = s.min()
    if pend_col:
        s = pd.to_datetime(work[pend_col], errors="coerce").dropna()
        if not s.empty: pmax = s.max()
    if pmin is not None: facts["period_min"] = pmin.strftime("%Y-%m-%d")
    if pmax is not None: facts["period_max"] = pmax.strftime("%Y-%m-%d")
    if pmin is not None and pmax is not None:
        facts["date_span_days"] = int((pmax - pmin).days)

    rmax = facts["rows_per_associate_max"]
    span = facts["date_span_days"]
    npd = facts["distinct_pay_dates"]

    if rmax <= 1:
        facts["detected_shape"] = "already_aggregated"
        facts["recommendation_reason"] = (
            "Each associate already has exactly one row; no aggregation needed."
        )
    elif span is not None and span >= 80 and npd >= 4:
        facts["detected_shape"] = "full_quarter_per_pay_period"
        facts["recommended_strategy"] = "full_quarter"
        facts["recommendation_reason"] = (
            f"Date span is {span} days with {npd} distinct pay dates and "
            f"{facts['rows_per_associate_avg']} rows per associate on average -- "
            f"this is a full-quarter per-pay-period export. "
            f"Recommend collapsing to one row per associate."
        )
    elif span is not None and span <= 40:
        facts["detected_shape"] = "partial_period"
        facts["recommended_strategy"] = "preserve_pay_periods"
        facts["recommendation_reason"] = (
            f"Date span is only {span} days -- this is a partial-period export. "
            f"Recommend preserving distinct pay periods (only merge same-day duplicates)."
        )
    else:
        facts["detected_shape"] = "ambiguous"
        span_txt = f"{span} days" if span is not None else "unknown"
        facts["recommendation_reason"] = (
            f"Date span ({span_txt}) is in-between full-quarter and partial. "
            f"Please choose 'full_quarter' or 'preserve_pay_periods' explicitly."
        )
    return facts


def _is_entry(v):
    """An 'entry' is any non-blank cell. 0 / 0.00 counts; blanks, '-', NaN don't."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    return str(v).strip() not in ("", "-", "nan", "NaT")


def find_retirement_columns(df):
    """Locate ALL 401k deferral columns and ALL Roth deferral columns
    (MEMO and LOAN columns excluded).

    Headers are matched on a lowercase alphanumeric-only normalization so
    'Voluntary K-401K', 'K-401K', '28-ADP 401K FLAT$' and '401(k)' all hit.
    Any header containing 'roth' goes to the Roth list. LOAN columns
    ('LN1-ADP 401K LOAN1') are never deferrals and are always skipped.
    Returns (k_cols, roth_cols) — both lists, possibly empty.
    """
    k_cols, roth_cols = [], []
    for c in df.columns:
        cl = str(c).strip().lower()
        if "memo" in cl or "loan" in cl:
            continue
        norm = re.sub(r"[^a-z0-9]", "", cl)
        if "roth" in norm:
            roth_cols.append(c)
        elif "401k" in norm:
            k_cols.append(c)
    return k_cols, roth_cols


def detect_memo_split(df):
    """Find the MEMO column carrying the combined 401k/Roth employer-match money.

    The target count is the number of rows holding an entry in ANY 401k
    deferral column OR ANY Roth column (loans excluded). A memo column whose
    own entry count equals the target is a match. Returns None when not
    applicable (no retirement columns or no memo columns), otherwise a dict:
      k_cols, roth_cols, target, memo_counts {col: count}, matches [cols]
    """
    k_cols, roth_cols = find_retirement_columns(df)
    if not k_cols and not roth_cols:
        return None
    memo_cols = [c for c in df.columns if "memo" in str(c).strip().lower()]
    if not memo_cols:
        return None

    union = pd.Series(False, index=df.index)
    for col in k_cols + roth_cols:
        union = union | df[col].map(_is_entry)
    target = int(union.sum())

    memo_counts = {c: int(df[c].map(_is_entry).sum()) for c in memo_cols}
    matches = [c for c, n in memo_counts.items() if n == target]
    return {
        "k_cols": k_cols,
        "roth_cols": roth_cols,
        "target": target,
        "memo_counts": memo_counts,
        "matches": matches,
    }


def split_memo_column(df, memo_col, k_cols, roth_cols):
    """Split the matched memo column into 401k vs Roth employer-match portions.

    Per row, the employee's deferrals decide where the match goes (loans are
    never counted; the 401k deferral is the SUM of all k_cols, e.g.
    K-ADP 401K + 28-ADP 401K FLAT$):

      1. 401k deferral only            -> entire match stays in `memo_col`
                                          (no split, even if match > deferral)
      2. Roth deferral only            -> entire match moves to `ROTH:<memo_col>`
      3. Both, match <= 401k deferral  -> entire match stays in `memo_col`
      4. Both, match >  401k deferral  -> keep up to the 401k deferral,
                                          move the excess to `ROTH:<memo_col>`
      -  Neither deferral present      -> match stays in `memo_col` and the
                                          row is flagged for review

    The new column is 'ROTH:<memo_col>' (UPPERCASE prefix — the UZIO prior
    payroll import uppercases file headers but matches mapping names
    case-sensitively, so the header must already be uppercase). Inserted
    immediately to the right of `memo_col`.

    Returns (new_df, {"new_col", "rows_split", "counts": {...},
                      "no_deferral": [employee ids]}).
    """
    df = df.copy()
    new_col = f"ROTH:{memo_col}"
    if new_col in df.columns:
        i = 1
        while f"{new_col}.{i}" in df.columns:
            i += 1
        new_col = f"{new_col}.{i}"

    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])

    def _row_sum(idx, cols):
        return sum((_to_float(df.at[idx, c]) or 0.0) for c in cols)

    kept_vals = []
    roth_vals = []
    rows_split = 0
    counts = {"kept_401k_only": 0, "moved_roth_only": 0,
              "kept_within_cap": 0, "split_excess": 0, "kept_no_deferral": 0}
    no_deferral = []

    for idx in df.index:
        memo_raw = df.at[idx, memo_col]
        memo_val = _to_float(memo_raw)
        if memo_val is None:
            kept_vals.append(memo_raw)
            roth_vals.append("")
            continue

        k_def = _row_sum(idx, k_cols)
        roth_def = _row_sum(idx, roth_cols)

        if k_def > 0 and roth_def <= 0:
            counts["kept_401k_only"] += 1
            kept_vals.append(memo_raw)
            roth_vals.append("")
        elif roth_def > 0 and k_def <= 0:
            counts["moved_roth_only"] += 1
            rows_split += 1
            kept_vals.append("")
            roth_vals.append(round(memo_val, 2))
        elif k_def > 0 and roth_def > 0:
            if memo_val <= k_def + 0.005:
                counts["kept_within_cap"] += 1
                kept_vals.append(memo_raw)
                roth_vals.append("")
            else:
                counts["split_excess"] += 1
                rows_split += 1
                roth_amt = round(memo_val - k_def, 2)
                kept = round(memo_val - roth_amt, 2)
                kept_vals.append(kept if kept != 0 else "")
                roth_vals.append(roth_amt)
        else:
            counts["kept_no_deferral"] += 1
            kept_vals.append(memo_raw)
            roth_vals.append("")
            if eid_col is not None:
                no_deferral.append(str(df.at[idx, eid_col]).strip())

    df[memo_col] = kept_vals
    df.insert(df.columns.get_loc(memo_col) + 1, new_col, roth_vals)
    return df, {"new_col": new_col, "rows_split": rows_split,
                "counts": counts, "no_deferral": no_deferral}


def render_ui():
    st.title("ADP - Prior Payroll Sanity Check")
    st.markdown(
        """
        Cleans an ADP Prior Payroll file so it can be ingested cleanly by downstream APIs.
        Three independent fix-ups are applied as needed:

        1. **Per-pay-period aggregation** -- when the implementor exported one row per
           pay period per employee, all rows for an Associate ID are folded into one
           (money/hours **summed**, period dates **min/max**'d, identity columns kept).
        2. **Duplicate row merge** -- two rows sharing the same Employee + Pay Date are
           smart-merged into one without double-counting.
        3. **Grand-total row removal** -- the bottom-of-file totals row where the
           previous employee's ID leaked is dropped.
        4. **401k / Roth memo split** -- when 401k / Roth deferrals exist (loans never
           counted), the MEMO column carrying the combined employer match is identified
           by entry count and split by each employee's own deferrals: 401k-only keeps
           everything, Roth-only moves entirely, both keeps up to the summed 401k
           deferral with only the excess moving to a new `ROTH:<memo column>` column.
        5. **Missing pay-period date columns** -- when `PERIOD BEGINNING DATE`,
           `PERIOD ENDING DATE`, or `PAY DATE` is absent (consolidated quarter files),
           the dates are read from the filename (`..._MMDDYYYY_MMDDYYYY_MMDDYYYY...`,
           begin / end / pay order) or entered manually, and the missing columns are
           inserted between `WORKED IN STATE` and `GROSS PAY` on every row.
        6. **Lived-in State / Local tax split** -- with the optional **Tax Validation
           Report** uploaded, single `LIVED-IN STATE / LOCAL - EMPLOYEE TAX` columns
           are split into per-jurisdiction columns (`LIVED-IN STATE (WI) - EMPLOYEE TAX`,
           `... (IL) ...`) by Associate ID, so each column maps to exactly one tax
           downstream.

        Upload an `.xlsx` / `.xls` / `.csv`. The cleaned output is **always a `.csv`** with
        the **exact same column headers and order** as the input (plus the new Roth
        memo column when the split applies).
        """
    )

    file = st.file_uploader(
        "Upload ADP Prior Payroll File",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="pps_input",
    )

    tax_val_file = st.file_uploader(
        "Tax Validation Report (optional — needed only to split LIVED-IN STATE / "
        "LIVED-IN LOCAL tax columns per jurisdiction)",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=False,
        key="pps_taxval",
        help="ADP Tax Validation Report with Associate ID, Lived in State Code and "
             "Lived in Local Jurisdiction columns. When the payroll file has a "
             "LIVED-IN STATE/LOCAL - EMPLOYEE TAX column, each employee's amount "
             "is moved into a per-jurisdiction column based on this report.",
    )

    if not file:
        st.info("Upload an ADP Prior Payroll file to begin.")
        return

    # ------------------------------------------------------------------
    # Step 1: Read + run detection BEFORE asking the user anything.
    # ------------------------------------------------------------------
    try:
        with st.spinner("Inspecting file..."):
            df_in, header_idx, sheet = read_input_file(file)
            original_count = len(df_in)
            df_a, summary_removed = drop_summary_rows(df_in)
            df_b, gt_info = detect_grand_total_row(df_a)
            date_found, date_missing = check_period_date_columns(df_b)
            fname_dates = parse_filename_dates(file.name) if date_missing else None
            date_fix_info = None
            if date_missing and fname_dates:
                df_b, date_fix_info = insert_period_date_columns(df_b, fname_dates, date_missing)
            facts = detect_file_shape(df_b)
    except Exception as e:
        st.error(f"Failed to read the file: {e}")
        return

    # ------------------------------------------------------------------
    # Step 2: Show the facts + recommendation. Always ask the user to
    # confirm; never apply silently.
    # ------------------------------------------------------------------
    st.markdown("### File shape detected")
    f1, f2, f3 = st.columns(3)
    f1.metric("Associates", facts["associates"])
    f2.metric("Total rows", facts["total_rows"])
    f3.metric("Rows / associate (max)", facts["rows_per_associate_max"])
    f4, f5, f6 = st.columns(3)
    f4.metric("Date span (days)", facts["date_span_days"] if facts["date_span_days"] is not None else "—")
    f5.metric("Distinct pay dates", facts["distinct_pay_dates"])
    f6.metric("Pay dates / associate (avg)", facts["distinct_pay_dates_per_associate_avg"])
    if facts["period_min"] and facts["period_max"]:
        st.caption(f"Pay period range: **{facts['period_min']} → {facts['period_max']}**")

    rec = facts["recommended_strategy"]
    if facts["detected_shape"] == "already_aggregated":
        st.info(
            "**Already aggregated.** Each associate already has exactly one row. "
            "Either strategy will leave the data as-is (just running grand-total / "
            "summary-row cleanup + the optional swap). Pick either one and run."
        )
        default_radio_idx = 0
    elif rec == "full_quarter":
        st.success(f"**Recommendation: Full Quarter.**  \n{facts['recommendation_reason']}")
        default_radio_idx = 0
    elif rec == "preserve_pay_periods":
        st.success(f"**Recommendation: Preserve Pay Periods.**  \n{facts['recommendation_reason']}")
        default_radio_idx = 1
    else:
        st.warning(f"**Recommendation: please choose explicitly.**  \n{facts['recommendation_reason']}")
        default_radio_idx = 0

    # ------------------------------------------------------------------
    # Step 2.2: Pay-period date columns (API requirement). Priority:
    # columns present -> untouched; missing -> filename dates; filename
    # unparseable -> manual date inputs (required before Run).
    # ------------------------------------------------------------------
    _PLACEMENT_TXT = {
        "between": "between `WORKED IN STATE` and `GROSS PAY`",
        "before_gross": "before `GROSS PAY` (no `WORKED IN STATE` column found)",
        "after_state": "after `WORKED IN STATE` (no `GROSS PAY` column found)",
        "appended": "at the end of the file (neither `WORKED IN STATE` nor `GROSS PAY` found)",
    }
    manual_dates = None
    if not date_missing:
        st.caption(
            "Pay-period date columns (PERIOD BEGINNING DATE / PERIOD ENDING DATE / "
            "PAY DATE) are already present -- left untouched."
        )
    else:
        st.markdown("### Pay-period date columns")
        missing_txt = ", ".join(f"`{c}`" for c in date_missing)
        if fname_dates:
            st.success(
                f"The file is missing {missing_txt}. Dates were read from the "
                f"filename and will be stamped on every row "
                f"({_PLACEMENT_TXT[date_fix_info['placement']]}):  \n"
                f"Period Beginning **{fname_dates['PERIOD BEGINNING DATE']}** · "
                f"Period Ending **{fname_dates['PERIOD ENDING DATE']}** · "
                f"Pay Date **{fname_dates['PAY DATE']}**.  \n"
                f"Please confirm these look right before running."
            )
        else:
            st.warning(
                f"The file is missing {missing_txt}, and the filename does not "
                f"contain three underscore-separated MMDDYYYY dates (expected e.g. "
                f"`PriorPayroll_01012026_01072026_01142026.xlsx`). Enter the dates "
                f"below -- they will be stamped on every row."
            )
            dcol1, dcol2, dcol3 = st.columns(3)
            d_begin = dcol1.date_input("Period Beginning Date", value=None,
                                       format="MM/DD/YYYY", key="pps_date_begin")
            d_end = dcol2.date_input("Period Ending Date", value=None,
                                     format="MM/DD/YYYY", key="pps_date_end")
            d_pay = dcol3.date_input("Pay Date", value=None,
                                     format="MM/DD/YYYY", key="pps_date_pay")
            if d_begin and d_end and d_pay:
                manual_dates = {
                    "PERIOD BEGINNING DATE": d_begin.strftime("%m/%d/%Y"),
                    "PERIOD ENDING DATE": d_end.strftime("%m/%d/%Y"),
                    "PAY DATE": d_pay.strftime("%m/%d/%Y"),
                }
                df_b, date_fix_info = insert_period_date_columns(df_b, manual_dates, date_missing)

    # ------------------------------------------------------------------
    # Step 2.3: Lived-in State / Local tax split. Runs BEFORE aggregation
    # so per-row values land in their jurisdiction column first. Only when
    # the payroll file has a lived-in column AND the Tax Validation Report
    # is uploaded; unmatched employees stay in the base column + flagged.
    # ------------------------------------------------------------------
    lived_cols = find_lived_in_columns(df_b)
    lived_split_infos = []
    lived_empty_dropped = []
    if lived_cols:
        st.markdown("### Lived-in State / Local tax split")
        money_counts = {k: int(df_b[c].map(_is_movable_value).sum())
                        for k, c in lived_cols.items()}
        # A lived-in column with no money serves no purpose — drop it outright,
        # report or not.
        for kind, col in list(lived_cols.items()):
            if not money_counts[kind]:
                df_b = df_b.drop(columns=[col])
                lived_empty_dropped.append(col)
                del lived_cols[kind]
        if lived_empty_dropped:
            st.caption(
                "Removed empty lived-in column(s): "
                + ", ".join(f"`{c}`" for c in lived_empty_dropped) + "."
            )
        if tax_val_file is None:
            if lived_cols:
                counts_txt = "; ".join(
                    f"`{c}` has money in {money_counts[k]} row(s)"
                    for k, c in lived_cols.items())
                st.warning(
                    f"This file has lived-in tax columns with money: {counts_txt}. "
                    f"**Upload the Tax Validation Report above** to split them into "
                    f"per-jurisdiction columns — without it the combined column "
                    f"cannot be mapped to multiple taxes downstream. "
                    f"(You can still run without splitting.)"
                )
        elif lived_cols:
            lookups, tv_err = load_tax_validation_report(tax_val_file)
            if tv_err:
                st.error(f"Tax Validation Report problem: {tv_err}")
            else:
                for kind, col in lived_cols.items():
                    df_b, sinfo = split_lived_in_column(df_b, col, lookups.get(kind, {}))
                    lived_split_infos.append((kind, col, sinfo))
                for kind, col, sinfo in lived_split_infos:
                    if sinfo["created"]:
                        base_txt = (
                            "the emptied base column was removed"
                            if sinfo["dropped_base"] else
                            "the base column was kept — it still holds unmatched value(s)"
                        )
                        st.success(
                            f"Split `{col}`: moved {sinfo['moved']} value(s) into "
                            + ", ".join(f"`{c}`" for c in sinfo["created"])
                            + f"; {base_txt}. The TAXABLE column is left combined."
                        )
                    if sinfo["unmatched"]:
                        st.warning(
                            f"**{len(sinfo['unmatched'])} row(s) in `{col}` could not be "
                            f"matched** — the Associate ID is missing from the Tax "
                            f"Validation Report (or its jurisdiction is blank there). "
                            f"Their values stay in the base column; please review."
                        )
                        st.dataframe(pd.DataFrame(sinfo["unmatched"]),
                                     hide_index=True, use_container_width=True)
    elif tax_val_file is not None:
        st.caption(
            "Tax Validation Report uploaded, but this payroll file has no "
            "LIVED-IN STATE / LIVED-IN LOCAL - EMPLOYEE TAX column — nothing to split."
        )

    st.markdown("### Confirm strategy and run")
    agg_choice = st.radio(
        "Aggregation Strategy (you can override the recommendation)",
        options=["Full Quarter — collapse to one row per associate",
                 "Preserve Pay Periods — keep each pay date, merge same-day duplicates only"],
        index=default_radio_idx,
        key="pps_agg_radio",
    )
    agg_strategy = "Full Quarter (Default)" if agg_choice.startswith("Full Quarter") else "Preserve Pay Periods"

    # ------------------------------------------------------------------
    # Step 2.5: Apply the chosen strategy in-memory so the 401k/Roth memo
    # detection runs on aggregated data ("aggregate first, then count"),
    # and let the user confirm the memo split before running.
    # ------------------------------------------------------------------
    try:
        with st.spinner("Applying strategy and analyzing memo columns..."):
            mode, period_info = detect_per_pay_period_structure(df_b)
            agg_info = None
            merge_events = None

            if mode == "aggregate":
                if agg_strategy == "Full Quarter (Default)":
                    df_c, agg_info = aggregate_by_associate(df_b)
                else:
                    df_c, merge_events = merge_duplicate_pay_periods(df_b)
                    mode = "preserve"
            else:
                df_c = df_b

            memo_info = detect_memo_split(df_c)
    except Exception as e:
        st.error(f"Failed to process the file: {e}")
        return

    chosen_memo = None
    if memo_info and memo_info["target"] > 0:
        k_cols = memo_info["k_cols"]
        roth_cols = memo_info["roth_cols"]
        matches = memo_info["matches"]
        ded_desc = " / ".join(f"`{c}`" for c in (k_cols + roth_cols))

        st.markdown("### 401k / Roth memo split")
        if not roth_cols:
            if matches:
                st.info(
                    f"Memo column `{matches[0]}` matches the {memo_info['target']} employees "
                    f"with {ded_desc}, but this file has no Roth column -- all memo money "
                    f"already belongs to 401k, so no split is needed."
                )
        elif len(matches) == 1:
            chosen_memo = matches[0]
            if k_cols:
                action_txt = (
                    f"each employee's match follows their own deferrals "
                    f"(loans never counted): 401k-only employees keep everything in "
                    f"`{chosen_memo}`; Roth-only employees move entirely to a new "
                    f"`ROTH:{chosen_memo}` column; employees with both keep up to their "
                    f"total 401k deferral ({' + '.join(f'`{c}`' for c in k_cols)}) and "
                    f"only the excess moves to Roth"
                )
            else:
                action_txt = (
                    f"there are no 401k deferral columns, so each Roth-deferring "
                    f"employee's match moves entirely to a new `ROTH:{chosen_memo}` column"
                )
            st.success(
                f"Matched memo column **`{chosen_memo}`** -- its entry count "
                f"({memo_info['target']}) equals the number of employees having "
                f"{ded_desc}. On run, {action_txt}."
            )
        elif len(matches) > 1:
            st.warning(
                f"**Multiple memo columns tie** at {memo_info['target']} entries "
                f"(= employees with {ded_desc}): "
                + ", ".join(f"`{c}`" for c in matches)
                + ". Pick the one carrying the 401k/Roth employer match, or skip the split."
            )
            pick = st.selectbox(
                "Memo column to split",
                options=["(Don't split)"] + matches,
                index=0,
                key="pps_memo_pick_tie",
            )
            chosen_memo = None if pick == "(Don't split)" else pick
        else:
            counts_txt = ", ".join(f"`{c}` = {n}" for c, n in memo_info["memo_counts"].items())
            st.warning(
                f"**No memo column matched.** {memo_info['target']} employees have "
                f"{ded_desc}, but the memo entry counts are: {counts_txt}. "
                f"You can still pick one manually, or skip the split."
            )
            label_map = {f"{c}  ({n} entries)": c for c, n in memo_info["memo_counts"].items()}
            pick = st.selectbox(
                "Memo column to split",
                options=["(Don't split)"] + list(label_map.keys()),
                index=0,
                key="pps_memo_pick_manual",
            )
            chosen_memo = label_map.get(pick)

    swap_net_take = st.checkbox(
        "Swap NET PAY and TAKE HOME values (the Carvan-style API expects them reversed)",
        value=True,
        key="pps_swap",
        help=(
            "When ON, the values in NET PAY and TAKE HOME are exchanged before download. "
            "Column headers stay the same -- only the data is swapped. "
            "Required for Carvan's API; uncheck if a client's API does not need it."
        ),
    )

    if not st.button("Run Sanity Check with this strategy", type="primary", use_container_width=True):
        return

    if date_missing and date_fix_info is None:
        st.error("Please fill in all three pay-period dates above before running.")
        return

    # ------------------------------------------------------------------
    # Step 3: Apply the confirmed memo split + optional swap.
    # ------------------------------------------------------------------
    try:
        with st.spinner("Finalizing..."):
            split_info = None
            if chosen_memo:
                df_c, split_info = split_memo_column(
                    df_c, chosen_memo, memo_info["k_cols"], memo_info["roth_cols"])

            swapped = False
            if swap_net_take:
                df_c, swapped = apply_net_take_swap(df_c)

            df_c = normalize_money_precision(df_c)
            final_count = len(df_c)
    except Exception as e:
        st.error(f"Failed to process the file: {e}")
        return

    st.success("Sanity check complete!")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Original Rows", original_count)
    m2.metric("Cleaned Rows", final_count)
    if mode == "aggregate":
        m3.metric("Mode", "Full Quarter Aggregation")
        m4.metric("Associates", agg_info["associates"] if agg_info else 0)
    elif mode == "preserve":
        m3.metric("Mode", "Preserved Pay Periods")
        m4.metric("Merged Dupes", len(merge_events) if merge_events else 0)
    else:
        m3.metric("Mode", "Already clean")
        m4.metric("Associates", period_info["associates"] if period_info else 0)

    note_lines = []
    if date_fix_info:
        date_src = "the filename" if fname_dates else "manual input"
        added_txt = ", ".join(f"`{c}`" for c in date_fix_info["added"])
        note_lines.append(
            f"Added missing date column(s) {added_txt} "
            f"{_PLACEMENT_TXT[date_fix_info['placement']]}, stamped on every row "
            f"from {date_src} (MM/DD/YYYY)."
        )
    if lived_empty_dropped:
        note_lines.append(
            "Removed empty lived-in column(s): "
            + ", ".join(f"`{c}`" for c in lived_empty_dropped) + "."
        )
    for _kind, _col, _sinfo in lived_split_infos:
        if _sinfo["created"]:
            base_txt = ("emptied base column removed" if _sinfo["dropped_base"]
                        else "base column kept (still holds unmatched values)")
            note_lines.append(
                f"Split `{_col}` per the Tax Validation Report: {_sinfo['moved']} value(s) "
                f"moved into {', '.join('`' + c + '`' for c in _sinfo['created'])}; {base_txt}."
            )
        if _sinfo["unmatched"]:
            note_lines.append(
                f"{len(_sinfo['unmatched'])} value(s) in `{_col}` left un-split — Associate ID "
                f"not in the Tax Validation Report (or jurisdiction blank). Review manually."
            )
    if summary_removed:
        note_lines.append(f"Dropped {summary_removed} interleaved 'Totals For Associate ID' summary rows from the raw file.")
    if gt_info:
        note_lines.append(
            f"Removed grand-total row carrying Employee ID `{gt_info['removed_employee_id']}` "
            f"({gt_info['removed_employee_name'] or 'name unknown'}). "
            f"Column `{gt_info['matched_on_column']}` held `{gt_info['matched_value']:,.2f}`, "
            f"about equal to the sum of preceding rows (`{gt_info['expected_sum']:,.2f}`)."
        )
    if mode == "aggregate" and period_info:
        max_pds = period_info.get("max_pay_dates_for_single_associate")
        max_msg = f", max {max_pds} pay dates for a single associate" if max_pds else ""
        note_lines.append(
            f"Detected per-pay-period file: {period_info['associates']} associates, "
            f"{period_info['with_multiple_rows']} with multiple rows{max_msg}. "
            f"Aggregated to one row per associate."
        )
    elif mode == "preserve" and period_info:
        note_lines.append(
            f"Preserved distinct pay periods. Successfully merged {len(merge_events)} "
            f"same-day duplicate row pairs." if merge_events else "Preserved distinct pay periods. No same-day duplicates found."
        )
    if split_info:
        c = split_info["counts"]
        stayed = c["kept_401k_only"] + c["kept_within_cap"]
        note_lines.append(
            f"Split memo column `{chosen_memo}` into `{split_info['new_col']}` by deferral: "
            f"{stayed} row(s) stayed as 401k match "
            f"({c['kept_401k_only']} 401k-only, {c['kept_within_cap']} within the 401k cap), "
            f"{c['moved_roth_only']} row(s) moved fully to Roth (Roth-only employees), "
            f"{c['split_excess']} row(s) split at the 401k-deferral cap. Loans were never counted."
        )
        if c["kept_no_deferral"]:
            ids = sorted(set(split_info["no_deferral"]))
            note_lines.append(
                f"{c['kept_no_deferral']} row(s) had match money but NO 401k or Roth deferral -- "
                f"kept as 401k match, please review: {', '.join(ids[:15])}"
                + (" ..." if len(ids) > 15 else "")
            )
    if swapped:
        note_lines.append("Swapped NET PAY and TAKE HOME values.")
    elif swap_net_take and not swapped:
        note_lines.append("Swap requested, but NET PAY and TAKE HOME columns were not found in the file.")
    if note_lines:
        st.warning("\n".join("- " + line for line in note_lines))

    if mode == "none" and not summary_removed and not gt_info and not split_info and not date_fix_info:
        st.info("No issues detected -- the cleaned output is identical to the input (minus formula evaluation).")

    with st.expander(f"Preview cleaned data ({final_count} rows)", expanded=False):
        st.dataframe(df_c.head(50), use_container_width=True)

    csv_buf = io.StringIO()
    df_c.to_csv(csv_buf, index=False)

    base_name = file.name.rsplit(".", 1)[0]
    st.download_button(
        label="Download Cleaned CSV",
        data=csv_buf.getvalue(),
        file_name=f"{base_name}_cleaned.csv",
        mime="text/csv",
        key="pps_download",
        use_container_width=True,
    )


if __name__ == "__main__":
    st.set_page_config(page_title="ADP Prior Payroll Sanity Check", layout="wide")
    render_ui()
