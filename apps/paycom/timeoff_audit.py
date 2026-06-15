import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "Paycom vs Uzio – Time Off Tool"

def clean_id(x):
    """Normalize Employee ID (remove .0, strip, remove leading zeros)."""
    if pd.isna(x): return ""
    s = str(x).strip()
    if s.endswith(".0"): s = s[:-2]
    # Remove leading zeros to match typically
    s = s.lstrip("0")
    return s

def run_tool(file_paycom, file_uzio):
    # 1. Read Paycom Report (Likely HTML disguised as XLS)
    try:
        # Try read_html first as it's common for Paycom 'xls'
        dfs = pd.read_html(file_paycom, header=0)
        if not dfs:
            st.error("No tables found in Paycom file.")
            return None
        df_p = dfs[0] # Assume main table is first
    except ValueError:
        # Fallback if actual Excel or CSV
        file_paycom.seek(0)
        try:
             df_p = pd.read_excel(file_paycom)
        except:
             file_paycom.seek(0)
             df_p = pd.read_csv(file_paycom)
    except Exception as e:
        st.error(f"Error reading Paycom file: {e}")
        return None

    # Normalize Paycom Columns
    # Look for 'Employee Code' and 'Net Available'
    p_cols = {c.strip(): c for c in df_p.columns}
    
    col_id_p = next((c for c in p_cols if "Employee Code" in c or "Employee ID" in c or "EECode" in c), None)
    col_bal_p = next((c for c in p_cols if "Net Available" in c), None)
    col_name_p = next((c for c in p_cols if "Employee Name" in c or "Name" in c or "Employee" in c), None)

    if not col_id_p or not col_bal_p:
        st.error(f"Could not find required columns in Paycom file. Found: {list(df_p.columns)}")
        return None

    # Create Lookup Map: CleanID -> Net Available
    # Filter out rows where Net Available is NaN/Blank if we want to preserve blanks?
    # Requirement: "Keep them as blank operating balance... dont fill anything"
    # So we only map values that exist.
    
    balance_map = {}
    for idx, row in df_p.iterrows():
        eid = clean_id(row[col_id_p])
        val = row[col_bal_p]
        
        if eid and pd.notna(val) and str(val).strip() != "":
            balance_map[eid] = val

    # ---------------------------------------------------------
    # PART A: Generate Clean Import File (using openpyxl)
    # ---------------------------------------------------------
    file_uzio.seek(0) # Reset file pointer for openpyxl
    try:
        wb = openpyxl.load_workbook(file_uzio)
    except Exception as e:
        st.error(f"Error reading Uzio Template with openpyxl: {e}")
        return None

    # Sheet 2 is index 1
    if len(wb.sheetnames) < 2:
        st.error("Uzio Template must have at least 2 sheets (Instruction, Time Off Details).")
        return None
        
    ws = wb.worksheets[1] # "Time Off Details"
    
    # Header is Row 4. Data starts Row 5.
    header_row = 4
    
    # Identify Columns in Header Row
    # openpyxl uses 1-based indexing for rows/cols
    # Iterate through header row to find column indices
    idx_id_u = None
    idx_bal_u = None
    
    for cell in ws[header_row]:
        val = str(cell.value).strip() if cell.value else ""
        if "Employee ID" in val:
            idx_id_u = cell.column # 1-based index
        elif "Operating Balance" in val or "Opening Balance" in val:
            idx_bal_u = cell.column
            
    if not idx_id_u or not idx_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening Balance' headers in Row 4 of Sheet 2.")
        return None

    # Iterate Data Rows
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=idx_id_u)
        cell_bal = ws.cell(row=row_idx, column=idx_bal_u)
        
        current_val = cell_bal.value
        
        # Rule: If Blank -> Keep Blank
        if current_val is None or str(current_val).strip() == "":
            continue # Skip
            
        # Policy Assigned -> Update
        eid = clean_id(cell_id.value)
        if eid in balance_map:
            cell_bal.value = balance_map[eid]
            
    # ---------------------------------------------------------
    # PART B: Calculate Audit Data (using pandas)
    # ---------------------------------------------------------
    file_uzio.seek(0) # Reset file pointer again for pandas
    # Re-read for pandas processing
    try:
        df_u = pd.read_excel(file_uzio, sheet_name=1, header=3)
    except Exception as e:
        st.error(f"Error reading Uzio Template for audit: {e}")
        return None
        
    u_cols = {c.strip(): c for c in df_u.columns}
    col_id_u = next((c for c in u_cols if "Employee ID" in c), None)
    col_bal_u = next((c for c in u_cols if "Operating Balance" in c), None)
    col_name_u = next((c for c in u_cols if "Employee Name" in c or "Name" in c), None)

    if not col_bal_u:
        col_bal_u = next((c for c in u_cols if "Opening Balance" in c), None)

    if not col_id_u or not col_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening/Operating Balance' in Uzio Template for audit. Found: {list(df_u.columns)}")
        return None

    # Future Columns
    col_future_app = next((c for c in p_cols if "Future Approved" in c), None)
    col_future_pend = next((c for c in p_cols if "Future Pending" in c), None)

    # Trackers
    matched_paycom_ids = set()
    unassigned_policies_rows = [] 
    
    # Exception List: [{'Employee ID', 'Employee Name', 'Issue Category', 'Paycom Balance', 'Future Approved', 'Future Pending'}]
    all_exceptions = []

    # function to apply map for audit tracking
    def audit_scan(row):
        current_val = row[col_bal_u]
        if pd.isna(current_val) or str(current_val).strip() == "":
            unassigned_policies_rows.append(row.to_dict())
            
            # Add to Exception Consolidated
            eid = str(row[col_id_u]) if pd.notna(row[col_id_u]) else ""
            name = str(row[col_name_u]) if col_name_u and pd.notna(row[col_name_u]) else "N/A"
            all_exceptions.append({
                'Employee ID': eid,
                'Employee Name': name,
                'Issue Category': 'Unassigned Policy (Blank Balance)',
                'Paycom Balance': '',
                'Future Approved': '',
                'Future Pending': ''
            })
            return
            
        eid = clean_id(row[col_id_u])
        if eid in balance_map:
            matched_paycom_ids.add(eid)

    df_u.apply(audit_scan, axis=1)

    # Additional Reports
    missing_in_uzio = []
    for idx, row in df_p.iterrows():
        eid = clean_id(row[col_id_p])
        val = row[col_bal_p]
        
        # Future Check (independent of missing)
        fa = float(row[col_future_app]) if col_future_app and pd.notna(row[col_future_app]) else 0
        fp = float(row[col_future_pend]) if col_future_pend and pd.notna(row[col_future_pend]) else 0
        
        if fa > 0 or fp > 0:
             name = str(row[col_name_p]) if col_name_p and pd.notna(row[col_name_p]) else "N/A"
             all_exceptions.append({
                'Employee ID': eid,
                'Employee Name': name,
                'Issue Category': 'Future Time Off Detected',
                'Paycom Balance': val,
                'Future Approved': fa,
                'Future Pending': fp
            })

        if eid and eid not in matched_paycom_ids:
            if pd.notna(val) and str(val).strip() != "":
                missing_in_uzio.append(row)
                
                # Add to Exception Consolidated
                name = str(row[col_name_p]) if col_name_p and pd.notna(row[col_name_p]) else "N/A"
                all_exceptions.append({
                    'Employee ID': eid,
                    'Employee Name': name,
                    'Issue Category': 'Missing in Uzio Template',
                    'Paycom Balance': val,
                    'Future Approved': fa,
                    'Future Pending': fp
                })
    
    df_missing = pd.DataFrame(missing_in_uzio)
    if df_missing.empty: df_missing = pd.DataFrame({'Message': ['All Paycom employees matched']})

    df_unassigned = pd.DataFrame(unassigned_policies_rows)
    if df_unassigned.empty: df_unassigned = pd.DataFrame({'Message': ['No unassigned policies found']})

    future_rows = []
    if col_future_app and col_future_pend:
        for idx, row in df_p.iterrows():
            try:
                fa = float(row[col_future_app]) if pd.notna(row[col_future_app]) else 0
                fp = float(row[col_future_pend]) if pd.notna(row[col_future_pend]) else 0
                if fa > 0 or fp > 0:
                    future_rows.append(row)
            except:
                pass
    df_future = pd.DataFrame(future_rows)
    if df_future.empty: df_future = pd.DataFrame({'Message': ['No future time off found']})

    # Consolidated Exceptions DataFrame
    df_exceptions = pd.DataFrame(all_exceptions)
    if df_exceptions.empty:
         df_exceptions = pd.DataFrame({'Message': ['No exceptions found']})

    # ---------------------------------------------------------
    # PART C: Append Audit Sheets to Workbook
    # ---------------------------------------------------------
    
    # Helper to add DF to Sheet
    def add_sheet_with_df(workbook, sheet_name, dataframe):
        ws_new = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws_new.append(r)

    add_sheet_with_df(wb, "Missing in Uzio", df_missing)
    add_sheet_with_df(wb, "Unassigned Policies", df_unassigned)
    add_sheet_with_df(wb, "Future Time Off", df_future)
    add_sheet_with_df(wb, "Paycom Raw Data", df_p)
    
    # Final Tab: Consolidated Exceptions in one glance
    add_sheet_with_df(wb, "Exception Summary", df_exceptions)

    out_final = io.BytesIO()
    wb.save(out_final)
    
    return out_final.getvalue()

def render_ui():
    st.title(APP_TITLE)
    client_name = st.text_input("Client Name", value="Client", key="paycom_timeoff_client")

    st.markdown("""
    **Instructions**:
    1. Upload **Paycom TimeOff Summary Report** (.xls / HTML).
    2. Upload **Uzio Time Off Import Template** (.xlsx).
    
    **Output**:
    - Generates a **Consolidated Excel File**.
    - **Original Tabs** (Instructions, Time Off Details) are preserved.
    - **Time Off Details** is updated with Paycom balances.
    - **New Tabs** added for Audit: Missing, Unassigned, Future, Raw Data.
    - *Note: Please delete the extra tabs before importing into Uzio.*
    """)

    col1, col2 = st.columns(2)
    with col1:
        f_p = st.file_uploader("Paycom TimeOff Report", type=["xls", "html", "xlsx"], key="pt_p")
    with col2:
        f_u = st.file_uploader("Uzio Template", type=["xlsx"], key="pt_u")

    if st.button("Generate Consolidated Report", key="run_timeoff"):
        if not f_u or not f_p:
            st.error("Please upload both files.")
            return
            
        try:
            with st.spinner("Processing..."):
                res = run_tool(f_p, f_u)
                
            if res:
                timestamp = pd.Timestamp.now().strftime('%d_%m_%Y_%H%M')
                filename = f"{client_name}_Uzio_Paycom_TimeOff_Audit_Report_{timestamp}.xlsx"

                st.success("File Generated Successfully!")
                st.download_button(
                    "Download Consolidated File",
                    data=res,
                    file_name=filename
                )
            else:
                st.error("No file could be generated due to an error.")

        except Exception as e:
            st.error(f"An error occurred: {e}")
            st.exception(e)

# Streamlit UI
if __name__ == "__main__":
    render_ui()
